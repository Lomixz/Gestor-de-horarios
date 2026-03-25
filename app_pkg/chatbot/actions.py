"""Execute chatbot tool actions against the database."""
import csv
import json
import logging
import secrets
import string
import threading
from collections import defaultdict
from io import BytesIO, StringIO
from datetime import datetime

from models import (
    db, User, Carrera, Materia, Grupo, Horario,
    HorarioAcademico, DisponibilidadProfesor, AsignacionProfesorGrupo,
    VersionHorario, BackupHistory, Role,
)
from app_pkg.helpers.schedule_helpers import procesar_horarios, obtener_periodo_actual
from app_pkg.helpers.generation_helpers import (
    generar_horarios_masivos_con_progreso,
    _generacion_progreso_fallback, _progreso_lock, _get_progress_tracker,
)
from . import file_manager

logger = logging.getLogger('sistema_academico')


def execute_tool(tool_name: str, arguments: dict, user: User) -> dict:
    """Execute a tool action and return the result.

    Returns:
        dict with keys: text (str), html (str|None), download_url (str|None)
    """
    handlers = {
        'get_schedule_group': _get_schedule_group,
        'get_schedule_professor': _get_schedule_professor,
        'get_my_schedule': _get_my_schedule,
        'list_groups': _list_groups,
        'list_professors': _list_professors,
        'list_subjects': _list_subjects,
        'get_stats': _get_stats,
        'add_user': _add_user,
        'create_backup': _create_backup,
        'generate_schedules': _generate_schedules,
        'set_availability': _set_availability,
        'get_professor_availability': _get_professor_availability,
        'get_my_subjects': _get_my_subjects,
        'get_my_groups': _get_my_groups,
        # New handlers
        'export_schedule_excel': _export_schedule_excel,
        'export_schedule_pdf': _export_schedule_pdf,
        'export_schedule_csv': _export_schedule_csv,
        'export_my_schedule': _export_my_schedule,
        'check_generation_progress': _check_generation_progress,
        'list_users': _list_users,
        'edit_user': _edit_user,
        'delete_user': _delete_user,
        'create_career': _create_career,
        'list_careers': _list_careers,
        'create_subject': _create_subject,
    }

    handler = handlers.get(tool_name)
    if not handler:
        return {'text': f'Herramienta "{tool_name}" no encontrada.', 'html': None, 'download_url': None}

    try:
        return handler(arguments, user)
    except Exception as e:
        logger.error(f"Error ejecutando tool {tool_name}: {e}", exc_info=True)
        return {'text': f'Error al ejecutar la acción: {str(e)}', 'html': None, 'download_url': None}


def _result(text, html=None, download_url=None, download_urls=None):
    return {'text': text, 'html': html, 'download_url': download_url, 'download_urls': download_urls}


# ---- Schedule queries ----

def _get_schedule_group(args, user):
    codigo = args.get('grupo_codigo', '').strip().upper()
    if not codigo:
        return _result('Se requiere el código del grupo.')

    grupo = Grupo.query.filter_by(codigo=codigo).first()
    if not grupo:
        return _result(f'No se encontró el grupo con código "{codigo}".')

    # Permission check for jefe_carrera
    if user.is_jefe_carrera() and not user.puede_acceder_carrera(grupo.carrera_id):
        return _result('No tienes permisos para ver este grupo.')

    horarios = HorarioAcademico.query.filter_by(grupo=codigo, activo=True).all()
    if not horarios:
        return _result(f'El grupo {codigo} no tiene horarios generados.')

    horarios.sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))

    # Build HTML table
    html = _build_schedule_table_group(horarios, codigo)

    # Generate Excel download
    excel_url = _generate_group_excel(horarios, codigo, user)
    downloads = []
    if excel_url:
        downloads.append({'url': excel_url, 'label': 'Excel'})

    # Build text representation so the LLM can describe the schedule
    text_lines = [f'Horario del grupo {codigo} ({grupo.carrera.nombre if grupo.carrera else ""}):']
    for h in horarios:
        dia = h.dia_semana.capitalize()
        hora = f'{h.get_hora_inicio_str()}-{h.get_hora_fin_str()}'
        materia = h.materia.nombre if h.materia else 'Sin materia'
        prof = h.profesor.get_nombre_completo() if h.profesor else 'Sin profesor'
        text_lines.append(f'- {dia} {hora}: {materia} ({prof})')

    return _result(
        '\n'.join(text_lines),
        html=html,
        download_urls=downloads if downloads else None,
    )


def _get_schedule_professor(args, user):
    nombre = args.get('nombre_profesor', '').strip()
    if not nombre:
        return _result('Se requiere el nombre del profesor.')

    # Search by name (partial match)
    profesores = User.query.filter(
        (User.nombre.ilike(f'%{nombre}%')) | (User.apellido.ilike(f'%{nombre}%')) |
        db.func.concat(User.nombre, ' ', User.apellido).ilike(f'%{nombre}%')
    ).filter(User.rol.in_(['profesor_completo', 'profesor_asignatura'])).all()

    if not profesores:
        return _result(f'No se encontró ningún profesor con el nombre "{nombre}".')

    if len(profesores) > 1:
        lista = '\n'.join([f"- {p.get_nombre_completo()}" for p in profesores[:10]])
        return _result(f'Se encontraron varios profesores. Por favor sé más específico:\n{lista}')

    profesor = profesores[0]

    # Permission check for jefe_carrera
    if user.is_jefe_carrera():
        carrera_ids = user.get_carreras_jefe_ids()
        prof_carreras = [c.id for c in profesor.carreras]
        if not any(c in carrera_ids for c in prof_carreras):
            return _result('No tienes permisos para ver el horario de este profesor.')

    horarios = HorarioAcademico.query.filter_by(profesor_id=profesor.id, activo=True).all()
    if not horarios:
        return _result(f'{profesor.get_nombre_completo()} no tiene horarios asignados.')

    horarios.sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))

    html = _build_schedule_table_professor(horarios, profesor)

    # Generate both Excel and PDF downloads
    downloads = _generate_professor_downloads(profesor, user)

    # Build text representation so the LLM can describe the schedule
    text_lines = [f'Horario de {profesor.get_nombre_completo()}:']
    for h in horarios:
        dia = h.dia_semana.capitalize()
        hora = f'{h.get_hora_inicio_str()}-{h.get_hora_fin_str()}'
        materia = h.materia.nombre if h.materia else 'Sin materia'
        text_lines.append(f'- {dia} {hora}: {materia} (Grupo: {h.grupo})')

    return _result(
        '\n'.join(text_lines),
        html=html,
        download_urls=downloads,
    )


def _get_my_schedule(args, user):
    horarios = HorarioAcademico.query.filter_by(profesor_id=user.id, activo=True).all()
    if not horarios:
        return _result('No tienes horarios asignados actualmente.')

    horarios.sort(key=lambda h: (h.get_dia_orden(), h.horario.hora_inicio))

    html = _build_schedule_table_professor(horarios, user)

    # Generate both Excel and PDF downloads
    downloads = _generate_professor_downloads(user, user)

    # Build text representation
    text_lines = ['Tu horario:']
    for h in horarios:
        dia = h.dia_semana.capitalize()
        hora = f'{h.get_hora_inicio_str()}-{h.get_hora_fin_str()}'
        materia = h.materia.nombre if h.materia else 'Sin materia'
        text_lines.append(f'- {dia} {hora}: {materia} (Grupo: {h.grupo})')

    return _result('\n'.join(text_lines), html=html, download_urls=downloads)


# ---- List queries ----

def _list_groups(args, user):
    carrera_filter = args.get('carrera', '').strip()
    query = Grupo.query

    if user.is_jefe_carrera():
        carrera_ids = user.get_carreras_jefe_ids()
        query = query.filter(Grupo.carrera_id.in_(carrera_ids))
    elif carrera_filter:
        carrera = Carrera.query.filter(
            (Carrera.codigo.ilike(f'%{carrera_filter}%')) | (Carrera.nombre.ilike(f'%{carrera_filter}%'))
        ).first()
        if carrera:
            query = query.filter_by(carrera_id=carrera.id)
        else:
            return _result(f'No se encontró la carrera "{carrera_filter}".')

    grupos = query.order_by(Grupo.codigo).all()
    if not grupos:
        return _result('No se encontraron grupos.')

    rows = []
    for g in grupos:
        carrera_nombre = g.carrera.nombre if g.carrera else '-'
        rows.append(f'<tr><td>{g.codigo}</td><td>{carrera_nombre}</td><td>{g.get_turno_display()}</td><td>{g.cuatrimestre}</td></tr>')

    html = f'''<table class="table table-sm table-striped">
<thead><tr><th>Código</th><th>Carrera</th><th>Turno</th><th>Cuatrimestre</th></tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>'''

    return _result(f'Se encontraron {len(grupos)} grupos:', html=html)


def _list_professors(args, user):
    carrera_filter = args.get('carrera', '').strip()
    query = User.query.filter(User.rol.in_(['profesor_completo', 'profesor_asignatura']))

    if carrera_filter:
        carrera = Carrera.query.filter(
            (Carrera.codigo.ilike(f'%{carrera_filter}%')) | (Carrera.nombre.ilike(f'%{carrera_filter}%'))
        ).first()
        if carrera:
            query = query.filter(User.carreras.any(Carrera.id == carrera.id))
        else:
            return _result(f'No se encontró la carrera "{carrera_filter}".')

    profesores = query.order_by(User.apellido, User.nombre).all()
    if not profesores:
        return _result('No se encontraron profesores.')

    rows = []
    for p in profesores:
        tipo = 'TC' if p.rol == 'profesor_completo' else 'PA'
        carreras = ', '.join([c.codigo for c in p.carreras]) if p.carreras else '-'
        rows.append(f'<tr><td>{p.get_nombre_completo()}</td><td>{tipo}</td><td>{carreras}</td></tr>')

    html = f'''<table class="table table-sm table-striped">
<thead><tr><th>Nombre</th><th>Tipo</th><th>Carreras</th></tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>'''

    return _result(f'Se encontraron {len(profesores)} profesores:', html=html)


def _list_subjects(args, user):
    carrera_filter = args.get('carrera', '').strip()
    query = Materia.query

    if carrera_filter:
        carrera = Carrera.query.filter(
            (Carrera.codigo.ilike(f'%{carrera_filter}%')) | (Carrera.nombre.ilike(f'%{carrera_filter}%'))
        ).first()
        if carrera:
            query = query.filter_by(carrera_id=carrera.id)
        else:
            return _result(f'No se encontró la carrera "{carrera_filter}".')

    materias = query.order_by(Materia.nombre).all()
    if not materias:
        return _result('No se encontraron materias.')

    rows = []
    for m in materias:
        carrera_nombre = m.carrera.codigo if m.carrera else '-'
        rows.append(f'<tr><td>{m.codigo}</td><td>{m.nombre}</td><td>{carrera_nombre}</td><td>{m.cuatrimestre}</td><td>{m.horas_semanales}</td></tr>')

    html = f'''<table class="table table-sm table-striped">
<thead><tr><th>Código</th><th>Nombre</th><th>Carrera</th><th>Cuatrimestre</th><th>Hrs/Sem</th></tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>'''

    return _result(f'Se encontraron {len(materias)} materias:', html=html)


# ---- Stats ----

def _get_stats(args, user):
    if user.is_jefe_carrera():
        carrera_ids = user.get_carreras_jefe_ids()
        grupos_count = Grupo.query.filter(Grupo.carrera_id.in_(carrera_ids)).count()
        materias_count = Materia.query.filter(Materia.carrera_id.in_(carrera_ids)).count()
        horarios_count = HorarioAcademico.query.filter(
            HorarioAcademico.activo == True,
            HorarioAcademico.grupo.in_(
                [g.codigo for g in Grupo.query.filter(Grupo.carrera_id.in_(carrera_ids)).all()]
            )
        ).count()
        carreras = Carrera.query.filter(Carrera.id.in_(carrera_ids)).all()
        carreras_nombres = ', '.join([c.nombre for c in carreras])

        text = f"""Estadísticas de tus carreras ({carreras_nombres}):
- Grupos: {grupos_count}
- Materias: {materias_count}
- Horarios activos: {horarios_count}"""
    else:
        users_count = User.query.count()
        carreras_count = Carrera.query.count()
        grupos_count = Grupo.query.count()
        materias_count = Materia.query.count()
        profesores_count = User.query.filter(User.rol.in_(['profesor_completo', 'profesor_asignatura'])).count()
        horarios_count = HorarioAcademico.query.filter_by(activo=True).count()

        text = f"""Estadísticas del sistema:
- Usuarios totales: {users_count}
- Carreras: {carreras_count}
- Grupos: {grupos_count}
- Materias: {materias_count}
- Profesores: {profesores_count}
- Horarios activos: {horarios_count}"""

    return _result(text)


# ---- Write actions ----

def _add_user(args, user):
    username = args.get('username', '').strip()
    password = args.get('password', '').strip()
    nombre = args.get('nombre', '').strip()
    apellido = args.get('apellido', '').strip()
    email = args.get('email', '').strip()
    rol = args.get('rol', '').strip()
    telefono = args.get('telefono', '').strip() or None
    tipo_profesor = args.get('tipo_profesor', '').strip() or None
    carreras_codigos = args.get('carreras', [])

    if not all([username, nombre, apellido, email, rol]):
        return _result('Se requieren username, nombre, apellido, email y rol para crear un usuario.')

    if rol not in ['admin', 'jefe_carrera', 'profesor_completo', 'profesor_asignatura']:
        return _result(f'Rol "{rol}" no válido. Opciones: admin, jefe_carrera, profesor_completo, profesor_asignatura')

    # Generate password if not provided
    if not password:
        alphabet = string.ascii_letters + string.digits + '!@#$%'
        password = ''.join(secrets.choice(alphabet) for _ in range(12))

    if User.query.filter_by(email=email).first():
        return _result(f'Ya existe un usuario con el email {email}.')

    if User.query.filter_by(username=username).first():
        return _result(f'Ya existe un usuario con el username {username}.')

    # Resolve carreras
    carreras_objs = []
    if carreras_codigos:
        for codigo in carreras_codigos:
            carrera = Carrera.query.filter_by(codigo=codigo.strip().upper()).first()
            if carrera:
                carreras_objs.append(carrera)
            else:
                return _result(f'No se encontró la carrera con código "{codigo}".')

    # Auto-set tipo_profesor based on rol
    if not tipo_profesor and rol in ('profesor_completo', 'profesor_asignatura'):
        tipo_profesor = 'completo' if rol == 'profesor_completo' else 'asignatura'

    new_user = User(
        username=username,
        email=email,
        password=password,
        nombre=nombre,
        apellido=apellido,
        rol=rol,
        telefono=telefono,
        tipo_profesor=tipo_profesor,
        requiere_cambio_password=True,
    )

    # Assign carreras
    for c in carreras_objs:
        new_user.carreras.append(c)

    db.session.add(new_user)

    # Assign role using add_role (same as UI — uses Role.get_or_create)
    new_user.add_role(rol)

    # Create default availability for professors
    if rol in ('profesor_completo', 'profesor_asignatura'):
        db.session.flush()  # Get new_user.id
        horarios = Horario.query.all()
        dias = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
        for dia in dias:
            for horario in horarios:
                disp = DisponibilidadProfesor(
                    profesor_id=new_user.id,
                    horario_id=horario.id,
                    dia_semana=dia,
                    disponible=True,
                    creado_por=user.id,
                )
                db.session.add(disp)

    db.session.commit()

    carreras_info = ', '.join([c.codigo for c in carreras_objs]) if carreras_objs else 'ninguna'
    return _result(
        f'Usuario creado exitosamente:\n'
        f'- Username: {username}\n'
        f'- Nombre: {nombre} {apellido}\n'
        f'- Email: {email}\n'
        f'- Rol: {rol}\n'
        f'- Carreras: {carreras_info}\n'
        f'- Password temporal: {password}\n'
        f'(El usuario deberá cambiar su contraseña en el primer inicio de sesión)'
    )


def _create_backup(args, user):
    try:
        import json as _json
        from datetime import datetime

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'backup_chatbot_{timestamp}.json'

        # Export all active schedules
        horarios = HorarioAcademico.query.filter_by(activo=True).all()
        datos = []
        for h in horarios:
            datos.append({
                'id': h.id, 'profesor_id': h.profesor_id,
                'materia_id': h.materia_id, 'horario_id': h.horario_id,
                'dia_semana': h.dia_semana, 'grupo': h.grupo,
                'periodo_academico': h.periodo_academico,
            })

        backup_data = _json.dumps(datos, ensure_ascii=False, indent=2)

        # Save to file manager
        file_id = file_manager.save_file(
            backup_data.encode('utf-8'),
            filename,
            user.id,
            'application/json'
        )

        # Record in backup history
        backup_record = BackupHistory(
            filename=filename,
            tipo='manual',
            tamano=len(backup_data),
            usuario_id=user.id,
        )
        db.session.add(backup_record)
        db.session.commit()

        return _result(
            f'Backup generado exitosamente: {filename} ({len(datos)} horarios)',
            download_url=f'/api/chatbot/download/{file_id}'
        )
    except Exception as e:
        return _result(f'Error al generar backup: {str(e)}')


def _generate_schedules(args, user):
    print(f"[CHATBOT] _generate_schedules called with args={args}", flush=True)
    version_nombre = args.get('version_nombre', '').strip()
    if not version_nombre:
        return _result(
            'Necesito un nombre para esta versión de horarios. '
            'Por ejemplo: "Horarios Enero 2026", "Versión Final", etc. '
            '¿Cómo quieres nombrar estos horarios?'
        )

    grupo_codigo = args.get('grupo_codigo', '').strip().upper()
    carrera_code = args.get('carrera', '').strip()

    # Case 1: Single specific group
    if grupo_codigo:
        grupo = Grupo.query.filter_by(codigo=grupo_codigo).first()
        if not grupo:
            return _result(f'No se encontró el grupo con código "{grupo_codigo}".')

        # Permission check for jefe_carrera
        if user.is_jefe_carrera() and not user.puede_acceder_carrera(grupo.carrera_id):
            return _result('No tienes permisos para generar horarios de este grupo.')

        return _generate_single_group(grupo, version_nombre, user)

    # Case 2: Multiple groups (by carrera or all)
    query = Grupo.query
    if user.is_jefe_carrera():
        carrera_ids = user.get_carreras_jefe_ids()
        query = query.filter(Grupo.carrera_id.in_(carrera_ids))
    elif carrera_code:
        carrera = Carrera.query.filter_by(codigo=carrera_code).first()
        if not carrera:
            return _result(f'No se encontró la carrera con código "{carrera_code}".')
        query = query.filter_by(carrera_id=carrera.id)

    grupos = query.all()
    if not grupos:
        return _result('No hay grupos para generar horarios.')

    grupos_ids = [g.id for g in grupos]
    return _generate_mass(grupos_ids, version_nombre, carrera_code, user)


def _generate_single_group(grupo, version_nombre, user):
    """Generate schedule for a single group using generar_horarios_secuencial (like the UI)."""
    from flask import current_app

    app = current_app._get_current_object()
    grupo_id = grupo.id
    grupo_codigo = grupo.codigo
    user_id = user.id

    año_actual = datetime.now().year
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    periodo_academico = f"{año_actual}-{año_actual}_{timestamp}"

    # Use Redis tracker (works across Gunicorn workers) with fallback
    tracker = _get_progress_tracker()
    task_id = None
    if tracker:
        task_id = tracker.create(total_grupos=1, iniciado_por=user_id)
        tracker.update(task_id,
                       codigo_grupo=grupo_codigo,
                       mensaje=f'Procesando grupo {grupo_codigo}...')
    else:
        with _progreso_lock:
            _generacion_progreso_fallback.update({
                'activo': True, 'total_grupos': 1, 'grupos_procesados': 0,
                'grupo_actual': grupo_codigo, 'codigo_grupo': grupo_codigo,
                'horarios_generados': 0, 'mensaje': f'Procesando grupo {grupo_codigo}...',
                'completado': False, 'exito': False, 'iniciado': True,
            })

    def _run():
        try:
            with app.app_context():
                from generador_horarios_mejorado import generar_horarios_secuencial
                import json as _json
                from collections import defaultdict

                resultado = generar_horarios_secuencial(
                    grupos_ids=[grupo_id],
                    periodo_academico=periodo_academico,
                    version_nombre=version_nombre,
                    creado_por=user_id,
                    dias_semana=['lunes', 'martes', 'miercoles', 'jueves', 'viernes'],
                )

                # 1. Backup FIRST (inline, same pattern as app.py:4078-4153)
                if resultado.get('exito') or resultado.get('horarios_generados', 0) > 0:
                    try:
                        grupos_por_carrera = defaultdict(list)
                        for gid in [grupo_id]:
                            g = Grupo.query.get(gid)
                            if g:
                                grupos_por_carrera[g.carrera_id].append(g.codigo.upper())

                        for carrera_id, grupos_codigos in grupos_por_carrera.items():
                            horarios_generados = HorarioAcademico.query.filter(
                                HorarioAcademico.grupo.in_(grupos_codigos),
                                HorarioAcademico.activo == True
                            ).all()

                            if not horarios_generados:
                                continue

                            datos = []
                            grupos_set = set()
                            for h in horarios_generados:
                                datos.append({
                                    'id': h.id, 'profesor_id': h.profesor_id,
                                    'materia_id': h.materia_id, 'horario_id': h.horario_id,
                                    'dia_semana': h.dia_semana, 'grupo': h.grupo,
                                    'periodo_academico': h.periodo_academico,
                                    'version_nombre': h.version_nombre
                                })
                                grupos_set.add(h.grupo)

                            carrera = Carrera.query.get(carrera_id)
                            carrera_nombre = carrera.nombre if carrera else 'Sin carrera'

                            version = VersionHorario(
                                nombre_version=version_nombre,
                                descripcion=f"Generación chatbot grupo {grupo_codigo} ({carrera_nombre})",
                                datos_horarios=_json.dumps(datos, ensure_ascii=False),
                                total_horarios=len(datos),
                                grupos_afectados=','.join(sorted(grupos_set)),
                                carrera_id=carrera_id,
                                creado_por=user_id,
                            )
                            db.session.add(version)
                            logger.info(f"Chatbot backup creado: {version_nombre} ({carrera_nombre}) con {len(datos)} horarios")

                        db.session.commit()

                        # Limit versions per user
                        usuario = User.query.get(user_id)
                        limite = 20 if usuario and usuario.is_admin() else 10
                        versiones = VersionHorario.query.filter(
                            VersionHorario.creado_por == user_id,
                            VersionHorario.activo == True
                        ).order_by(VersionHorario.fecha_creacion.desc()).all()
                        if len(versiones) > limite:
                            for v in versiones[limite:]:
                                v.activo = False
                            db.session.commit()

                    except Exception as e:
                        logger.error(f"Error creando backup chatbot: {e}", exc_info=True)

                # 2. Mark complete AFTER backup
                if tracker and task_id:
                    tracker.complete(task_id, resultado.get('exito', False), resultado.get('mensaje', ''))
                else:
                    with _progreso_lock:
                        _generacion_progreso_fallback.update({
                            'activo': False, 'completado': True,
                            'exito': resultado.get('exito', False),
                            'grupos_procesados': resultado.get('grupos_procesados', 0),
                            'horarios_generados': resultado.get('horarios_generados', 0),
                            'mensaje': resultado.get('mensaje', ''),
                        })

        except Exception as e:
            logger.error(f"Error en thread de generación single-group: {e}", exc_info=True)
            if tracker and task_id:
                tracker.complete(task_id, False, f'Error: {str(e)}')
            else:
                with _progreso_lock:
                    _generacion_progreso_fallback.update({
                        'activo': False, 'completado': True, 'exito': False,
                        'mensaje': f'Error: {str(e)}',
                    })

    thread = threading.Thread(target=_run)
    thread.daemon = True
    thread.start()

    return _result(
        f'Generando horario del grupo {grupo_codigo} con el nombre "{version_nombre}". '
        f'Pregúntame "¿cómo va la generación?" para ver el progreso.'
    )


def _generate_mass(grupos_ids, version_nombre, carrera_code, user):
    """Generate schedules for multiple groups with progress tracking and backup."""
    from flask import current_app

    app = current_app._get_current_object()
    user_id = user.id
    periodo, _ = obtener_periodo_actual()

    # Use Redis tracker (works across Gunicorn workers) with fallback
    tracker = _get_progress_tracker()
    task_id = None
    if tracker:
        task_id = tracker.create(total_grupos=len(grupos_ids), iniciado_por=user_id)
    else:
        with _progreso_lock:
            _generacion_progreso_fallback.clear()
            _generacion_progreso_fallback.update({
                'activo': True, 'total_grupos': len(grupos_ids), 'grupos_procesados': 0,
                'grupo_actual': 'Iniciando...', 'codigo_grupo': '',
                'horarios_generados': 0, 'mensaje': 'Iniciando generación...',
                'completado': False, 'exito': False, 'iniciado': True,
            })

    def _run():
        try:
            with app.app_context():
                import json as _json
                from collections import defaultdict

                resultado = generar_horarios_masivos_con_progreso(
                    grupos_ids=grupos_ids,
                    periodo_academico=periodo,
                    version_nombre=version_nombre,
                    creado_por=user_id,
                    dias_semana=['lunes', 'martes', 'miercoles', 'jueves', 'viernes'],
                    task_id=task_id,
                )

                # 1. Backup FIRST (inline, exact pattern from app.py:4078-4153)
                logger.info(f"Chatbot generación masiva completada: exito={resultado.get('exito')}, horarios={resultado.get('horarios_generados', 0)}")
                if resultado.get('exito') or resultado.get('horarios_generados', 0) > 0:
                    try:
                        grupos_por_carrera = defaultdict(list)
                        for gid in grupos_ids:
                            grupo = Grupo.query.get(gid)
                            if grupo:
                                grupos_por_carrera[grupo.carrera_id].append(grupo.codigo.upper())

                        logger.info(f"Chatbot creando backups para {len(grupos_por_carrera)} carrera(s)")

                        for carrera_id, grupos_codigos in grupos_por_carrera.items():
                            horarios_generados = HorarioAcademico.query.filter(
                                HorarioAcademico.grupo.in_(grupos_codigos),
                                HorarioAcademico.activo == True
                            ).all()

                            if not horarios_generados:
                                continue

                            datos = []
                            grupos_set = set()
                            for h in horarios_generados:
                                datos.append({
                                    'id': h.id, 'profesor_id': h.profesor_id,
                                    'materia_id': h.materia_id, 'horario_id': h.horario_id,
                                    'dia_semana': h.dia_semana, 'grupo': h.grupo,
                                    'periodo_academico': h.periodo_academico,
                                    'version_nombre': h.version_nombre
                                })
                                grupos_set.add(h.grupo)

                            carrera = Carrera.query.get(carrera_id)
                            carrera_nombre = carrera.nombre if carrera else 'Sin carrera'

                            version = VersionHorario(
                                nombre_version=version_nombre,
                                descripcion=f"Generación masiva chatbot de {len(grupos_codigos)} grupos ({carrera_nombre})",
                                datos_horarios=_json.dumps(datos, ensure_ascii=False),
                                total_horarios=len(datos),
                                grupos_afectados=','.join(sorted(grupos_set)),
                                carrera_id=carrera_id,
                                creado_por=user_id,
                            )
                            db.session.add(version)
                            logger.info(f"Chatbot backup creado: {version_nombre} ({carrera_nombre}) con {len(datos)} horarios")

                        db.session.commit()

                        # Limit versions per user — 20 for admin, 10 for others
                        usuario = User.query.get(user_id)
                        limite = 20 if usuario and usuario.is_admin() else 10
                        versiones = VersionHorario.query.filter(
                            VersionHorario.creado_por == user_id,
                            VersionHorario.activo == True
                        ).order_by(VersionHorario.fecha_creacion.desc()).all()
                        if len(versiones) > limite:
                            for v in versiones[limite:]:
                                v.activo = False
                            db.session.commit()

                    except Exception as e:
                        logger.error(f"Error creando backup chatbot: {e}", exc_info=True)
                else:
                    logger.info(f"Chatbot backup saltado: exito={resultado.get('exito')}, horarios={resultado.get('horarios_generados', 0)}")

                # 2. Mark complete AFTER backup
                if tracker and task_id:
                    tracker.complete(task_id, resultado.get('exito', False), resultado.get('mensaje', ''))
                else:
                    with _progreso_lock:
                        _generacion_progreso_fallback.update({
                            'activo': False, 'completado': True,
                            'exito': resultado.get('exito', False),
                            'grupos_procesados': resultado.get('grupos_procesados', 0),
                            'horarios_generados': resultado.get('horarios_generados', 0),
                            'mensaje': resultado.get('mensaje', ''),
                        })

        except Exception as e:
            logger.error(f"Error en thread de generación masiva chatbot: {e}", exc_info=True)
            if tracker and task_id:
                tracker.complete(task_id, False, f'Error: {str(e)}')
            else:
                with _progreso_lock:
                    _generacion_progreso_fallback.update({
                        'activo': False, 'completado': True, 'exito': False,
                        'mensaje': f'Error: {str(e)}',
                    })

    thread = threading.Thread(target=_run)
    thread.daemon = True
    thread.start()

    carreras_info = ''
    if carrera_code:
        carreras_info = f' de la carrera {carrera_code}'
    elif user.is_jefe_carrera():
        carreras_info = ' de tus carreras'

    return _result(
        f'Iniciando generación de horarios{carreras_info} para {len(grupos_ids)} grupos '
        f'con el nombre "{version_nombre}". '
        f'Este proceso puede tardar varios minutos. '
        f'Pregúntame "¿cómo va la generación?" para ver el progreso.'
    )


def _set_availability(args, user):
    dia = args.get('dia', '').strip().lower()
    horario_nombre = args.get('horario_nombre', '').strip()
    disponible = args.get('disponible', True)

    if not dia or not horario_nombre:
        return _result('Se requiere el día y el nombre del horario.')

    dias_validos = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
    if dia not in dias_validos:
        return _result(f'Día no válido. Opciones: {", ".join(dias_validos)}')

    horario = Horario.query.filter(Horario.nombre.ilike(f'%{horario_nombre}%')).first()
    if not horario:
        horarios_disponibles = [h.nombre for h in Horario.query.order_by(Horario.orden).all()]
        return _result(f'Horario no encontrado. Opciones disponibles: {", ".join(horarios_disponibles)}')

    disp = DisponibilidadProfesor.query.filter_by(
        profesor_id=user.id, horario_id=horario.id, dia_semana=dia
    ).first()

    if disp:
        disp.disponible = disponible
    else:
        disp = DisponibilidadProfesor(
            profesor_id=user.id,
            horario_id=horario.id,
            dia_semana=dia,
            disponible=disponible
        )
        db.session.add(disp)

    db.session.commit()

    estado = 'disponible' if disponible else 'no disponible'
    return _result(f'Disponibilidad actualizada: {dia.capitalize()} {horario.nombre} → {estado}')


def _get_professor_availability(args, user):
    """Query professor availability (configured + real free hours).

    Uses the same approach as the admin view (hora_inicio.hour as grid key)
    to ensure data matches exactly.
    """
    nombre = args.get('nombre_profesor', '').strip()
    if not nombre:
        return _result('Se requiere el nombre del profesor.')

    # Search professor (reuse same logic as _get_schedule_professor)
    profesores = User.query.filter(
        (User.nombre.ilike(f'%{nombre}%')) | (User.apellido.ilike(f'%{nombre}%')) |
        db.func.concat(User.nombre, ' ', User.apellido).ilike(f'%{nombre}%')
    ).filter(User.rol.in_(['profesor_completo', 'profesor_asignatura'])).all()

    if not profesores:
        return _result(f'No se encontró ningún profesor con el nombre "{nombre}".')

    if len(profesores) > 1:
        lista = '\n'.join([f"- {p.get_nombre_completo()}" for p in profesores[:10]])
        return _result(f'Se encontraron varios profesores. Por favor sé más específico:\n{lista}')

    profesor = profesores[0]

    # Permission check for jefe_carrera
    if user.is_jefe_carrera():
        carrera_ids = user.get_carreras_jefe_ids()
        prof_carreras = [c.id for c in profesor.carreras]
        if not any(c in carrera_ids for c in prof_carreras):
            return _result('No tienes permisos para ver la disponibilidad de este profesor.')

    dias = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes']
    dias_display = {'lunes': 'Lun', 'martes': 'Mar', 'miercoles': 'Mié', 'jueves': 'Jue', 'viernes': 'Vie'}

    # Get actual time slots from the Gestión de Horarios module (not a fixed range)
    horarios_sistema = Horario.query.filter_by(activo=True).order_by(Horario.hora_inicio).all()
    if not horarios_sistema:
        return _result('No hay horarios configurados en el sistema.')

    # Build unique time slots sorted by hora_inicio, merging matutino/vespertino
    # Each slot maps to the list of Horario IDs that share that time
    slots_unicos = {}  # key: (hora_inicio, hora_fin) -> [horario_id, ...]
    for h in horarios_sistema:
        slot_key = (h.hora_inicio, h.hora_fin)
        if slot_key not in slots_unicos:
            slots_unicos[slot_key] = []
        slots_unicos[slot_key].append(h.id)
    # Sort by hora_inicio
    slots_ordenados = sorted(slots_unicos.items(), key=lambda x: x[0][0])

    # Build class map keyed by (horario_id, dia)
    clases = HorarioAcademico.query.filter_by(profesor_id=profesor.id, activo=True).all()
    clases_por_hid = {}  # key: (horario_id, dia) -> "materia (grupo)"
    for c in clases:
        if c.horario and c.materia:
            clases_por_hid[(c.horario_id, c.dia_semana.lower())] = f'{c.materia.nombre} ({c.grupo})'

    # Build availability map keyed by (horario_id, dia)
    disponibilidades = DisponibilidadProfesor.query.filter_by(
        profesor_id=profesor.id, activo=True
    ).all()
    disp_por_hid = {}  # key: (horario_id, dia) -> bool
    for d in disponibilidades:
        disp_por_hid[(d.horario_id, d.dia_semana.lower())] = d.disponible

    # Build table using actual system time slots
    rows = []
    total_clase = 0
    total_libre = 0
    total_no_disp = 0

    for (h_inicio, h_fin), horario_ids in slots_ordenados:
        slot_label = f'{h_inicio.strftime("%H:%M")}-{h_fin.strftime("%H:%M")}'
        cells = f'<td><strong>{slot_label}</strong></td>'

        for dia in dias:
            # Check ALL horario IDs for this time slot (handles matutino/vespertino)
            clases_en_celda = []
            es_disponible = None
            for hid in horario_ids:
                key = (hid, dia)
                if key in clases_por_hid:
                    clases_en_celda.append(clases_por_hid[key])
                if key in disp_por_hid:
                    if disp_por_hid[key] is False:
                        es_disponible = False
                    elif es_disponible is None:
                        es_disponible = disp_por_hid[key]

            if clases_en_celda:
                clases_html = '<br>'.join(clases_en_celda)
                cells += f'<td style="background:#fee2e2;color:#991b1b;font-size:9px">{clases_html}</td>'
                total_clase += len(clases_en_celda)
            elif es_disponible is False:
                cells += '<td style="background:#fef3c7;color:#92400e">No disp.</td>'
                total_no_disp += 1
            elif es_disponible is True:
                cells += '<td style="background:#d1fae5;color:#065f46">Libre</td>'
                total_libre += 1
            else:
                cells += '<td style="background:#f3f4f6;color:#6b7280">Sin config.</td>'
                total_libre += 1

        rows.append(f'<tr>{cells}</tr>')

    headers = ''.join([f'<th>{dias_display[d]}</th>' for d in dias])
    html = f'''<table class="table table-sm table-bordered">
<thead class="table-dark"><tr><th>Hora</th>{headers}</tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>
<div style="font-size:9px;margin-top:4px">
<span style="background:#d1fae5;padding:1px 4px">Libre</span>
<span style="background:#fee2e2;padding:1px 4px">Con clase</span>
<span style="background:#fef3c7;padding:1px 4px">No disponible</span>
<span style="background:#f3f4f6;padding:1px 4px">Sin configurar</span>
</div>'''

    text = (f'Disponibilidad de {profesor.get_nombre_completo()}:\n'
            f'- Horas con clase: {total_clase}\n'
            f'- Horas libres: {total_libre}\n'
            f'- Horas no disponibles: {total_no_disp}')

    return _result(text, html=html)


def _get_my_subjects(args, user):
    materias = user.materias if hasattr(user, 'materias') else []
    if not materias:
        return _result('No tienes materias asignadas.')

    rows = []
    for m in materias:
        carrera = m.carrera.codigo if m.carrera else '-'
        rows.append(f'<tr><td>{m.codigo}</td><td>{m.nombre}</td><td>{carrera}</td><td>{m.horas_semanales}</td></tr>')

    html = f'''<table class="table table-sm table-striped">
<thead><tr><th>Código</th><th>Nombre</th><th>Carrera</th><th>Hrs/Sem</th></tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>'''

    return _result(f'Tienes {len(materias)} materias asignadas:', html=html)


def _get_my_groups(args, user):
    asignaciones = AsignacionProfesorGrupo.query.filter_by(profesor_id=user.id).all()
    if not asignaciones:
        return _result('No tienes grupos asignados.')

    grupos_ids = set(a.grupo_id for a in asignaciones)
    grupos = Grupo.query.filter(Grupo.id.in_(grupos_ids)).all()

    rows = []
    for g in grupos:
        carrera = g.carrera.nombre if g.carrera else '-'
        rows.append(f'<tr><td>{g.codigo}</td><td>{carrera}</td><td>{g.get_turno_display()}</td></tr>')

    html = f'''<table class="table table-sm table-striped">
<thead><tr><th>Código</th><th>Carrera</th><th>Turno</th></tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>'''

    return _result(f'Tienes {len(grupos)} grupos asignados:', html=html)


# ---- Export tools ----

def _export_schedule_excel(args, user):
    tipo = args.get('tipo', '').strip().lower()
    identificador = args.get('identificador', '').strip()

    if not tipo or not identificador:
        return _result('Se requiere el tipo (grupo/profesor) y el identificador.')

    if tipo == 'grupo':
        codigo = identificador.upper()
        grupo = Grupo.query.filter_by(codigo=codigo).first()
        if not grupo:
            return _result(f'No se encontró el grupo "{codigo}".')

        horarios = HorarioAcademico.query.filter_by(grupo=codigo, activo=True).all()
        if not horarios:
            return _result(f'El grupo {codigo} no tiene horarios generados.')

        download_url = _generate_group_excel(horarios, codigo, user)
        if not download_url:
            return _result('Error al generar el archivo Excel.')
        return _result(f'Excel del horario del grupo {codigo} generado:', download_url=download_url)

    elif tipo == 'profesor':
        profesor = _find_single_professor(identificador)
        if isinstance(profesor, dict):
            return profesor  # Error result

        download_url = _generate_professor_excel(profesor, user)
        if not download_url:
            return _result('Error al generar el archivo Excel. El profesor puede no tener horarios.')
        return _result(f'Excel del horario de {profesor.get_nombre_completo()} generado:', download_url=download_url)

    return _result('Tipo no válido. Usa "grupo" o "profesor".')


def _export_schedule_pdf(args, user):
    nombre = args.get('nombre_profesor', '').strip()
    if not nombre:
        return _result('Se requiere el nombre del profesor.')

    profesor = _find_single_professor(nombre)
    if isinstance(profesor, dict):
        return profesor  # Error result

    try:
        from app_pkg.helpers.export_helpers import generar_pdf_profesor_buffer

        buf, filename = generar_pdf_profesor_buffer(profesor.get_nombre_completo())
        if buf is None:
            return _result(f'Error al generar PDF: {filename}')

        file_id = file_manager.save_file(
            buf.getvalue(),
            filename,
            user.id,
            'application/pdf'
        )
        return _result(
            f'PDF del horario de {profesor.get_nombre_completo()} generado:',
            download_url=f'/api/chatbot/download/{file_id}'
        )
    except Exception as e:
        logger.error(f"Error generando PDF de profesor: {e}", exc_info=True)
        return _result(f'Error al generar PDF: {str(e)}')


def _export_my_schedule(args, user):
    formato = args.get('formato', 'excel').strip().lower()

    if formato == 'excel':
        download_url = _generate_professor_excel(user, user)
        if not download_url:
            return _result('No se pudo generar el Excel. Puede que no tengas horarios asignados.')
        return _result('Tu horario en Excel:', download_url=download_url)

    elif formato == 'pdf':
        try:
            from app_pkg.helpers.export_helpers import generar_pdf_profesor_buffer

            buf, filename = generar_pdf_profesor_buffer(user.get_nombre_completo())
            if buf is None:
                return _result(f'Error al generar PDF: {filename}')

            file_id = file_manager.save_file(
                buf.getvalue(),
                filename,
                user.id,
                'application/pdf'
            )
            return _result('Tu horario en PDF:', download_url=f'/api/chatbot/download/{file_id}')
        except Exception as e:
            logger.error(f"Error generando PDF propio: {e}", exc_info=True)
            return _result(f'Error al generar PDF: {str(e)}')

    elif formato == 'csv':
        download_url = _generate_professor_csv(user, user)
        if not download_url:
            return _result('No se pudo generar el CSV. Puede que no tengas horarios asignados.')
        return _result('Tu horario en CSV:', download_url=download_url)

    return _result('Formato no válido. Usa "excel", "pdf" o "csv".')


def _export_schedule_csv(args, user):
    """Export a group or professor schedule as CSV."""
    tipo = args.get('tipo', '').strip().lower()
    identificador = args.get('identificador', '').strip()

    if not tipo or not identificador:
        return _result('Se requiere el tipo (grupo/profesor) y el identificador.')

    if tipo == 'grupo':
        codigo = identificador.upper()
        grupo = Grupo.query.filter_by(codigo=codigo).first()
        if not grupo:
            return _result(f'No se encontró el grupo "{codigo}".')

        horarios = HorarioAcademico.query.filter_by(grupo=codigo, activo=True).all()
        if not horarios:
            return _result(f'El grupo {codigo} no tiene horarios generados.')

        download_url = _generate_group_csv(horarios, codigo, user)
        if not download_url:
            return _result('Error al generar el archivo CSV.')
        return _result(f'CSV del horario del grupo {codigo} generado:', download_url=download_url)

    elif tipo == 'profesor':
        profesor = _find_single_professor(identificador)
        if isinstance(profesor, dict):
            return profesor

        download_url = _generate_professor_csv(profesor, user)
        if not download_url:
            return _result('Error al generar el archivo CSV. El profesor puede no tener horarios.')
        return _result(f'CSV del horario de {profesor.get_nombre_completo()} generado:', download_url=download_url)

    return _result('Tipo no válido. Usa "grupo" o "profesor".')


# ---- Progress check ----

def _check_generation_progress(args, user):
    tracker = _get_progress_tracker()

    if tracker:
        # Try to find active/recent tasks in Redis
        try:
            active_tasks = tracker.get_active_tasks()
            if active_tasks:
                # Prefer tasks that are still active, otherwise show most recent completed
                active = [t for t in active_tasks if t.get('activo') and not t.get('completado')]
                task = active[0] if active else active_tasks[0]

                total = task.get('total_grupos', 0)
                procesados = task.get('grupos_procesados', 0)
                grupo_actual = task.get('codigo_grupo', task.get('grupo_actual', ''))
                completado = task.get('completado', False)
                mensaje = task.get('mensaje', '')
                horarios = task.get('horarios_generados', 0)

                if completado:
                    exito = task.get('exito', False)
                    estado = 'exitosamente' if exito else 'con errores'
                    # Clean up completed task from Redis
                    task_id = task.get('task_id')
                    if task_id:
                        tracker.delete(task_id)
                    return _result(
                        f'La generación ha finalizado {estado}.\n'
                        f'- Grupos procesados: {procesados}/{total}\n'
                        f'- Horarios generados: {horarios}\n'
                        f'- {mensaje}'
                    )

                porcentaje = (procesados / total * 100) if total > 0 else 0
                return _result(
                    f'Generación en curso:\n'
                    f'- Grupos procesados: {procesados}/{total} ({porcentaje:.0f}%)\n'
                    f'- Grupo actual: {grupo_actual}\n'
                    f'- {mensaje}'
                )
        except Exception as e:
            logger.error(f"Error consultando progreso en Redis: {e}")

    # Fallback to global dict (single-worker / dev mode)
    with _progreso_lock:
        prog = dict(_generacion_progreso_fallback)

    if not prog.get('activo') and not prog.get('completado') and not prog.get('iniciado'):
        return _result('No hay ninguna generación de horarios en curso.')

    total = prog.get('total_grupos', 0)
    procesados = prog.get('grupos_procesados', 0)
    grupo_actual = prog.get('grupo_actual', '')
    completado = prog.get('completado', False)
    mensaje = prog.get('mensaje', '')

    if completado:
        exito = prog.get('exito', False)
        estado = 'exitosamente' if exito else 'con errores'
        return _result(
            f'La generación ha finalizado {estado}.\n'
            f'- Grupos procesados: {procesados}/{total}\n'
            f'- Horarios generados: {prog.get("horarios_generados", 0)}\n'
            f'- {mensaje}'
        )

    porcentaje = (procesados / total * 100) if total > 0 else 0
    return _result(
        f'Generación en curso:\n'
        f'- Grupos procesados: {procesados}/{total} ({porcentaje:.0f}%)\n'
        f'- Grupo actual: {grupo_actual}\n'
        f'- {mensaje}'
    )


# ---- Admin CRUD tools ----

def _list_users(args, user):
    query = User.query
    rol_filter = args.get('rol', '').strip() if args.get('rol') else ''
    activo_filter = args.get('activo')

    if rol_filter:
        query = query.filter_by(rol=rol_filter)
    if activo_filter is not None:
        query = query.filter_by(activo=activo_filter)

    usuarios = query.order_by(User.apellido, User.nombre).all()
    if not usuarios:
        return _result('No se encontraron usuarios con esos criterios.')

    rows = []
    for u in usuarios:
        estado = 'Activo' if u.activo else 'Inactivo'
        carreras = ', '.join([c.codigo for c in u.carreras]) if u.carreras else '-'
        rows.append(
            f'<tr><td>{u.username}</td><td>{u.get_nombre_completo()}</td>'
            f'<td>{u.email}</td><td>{u.rol}</td><td>{carreras}</td><td>{estado}</td></tr>'
        )

    html = f'''<table class="table table-sm table-striped">
<thead><tr><th>Username</th><th>Nombre</th><th>Email</th><th>Rol</th><th>Carreras</th><th>Estado</th></tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>'''

    return _result(f'Se encontraron {len(usuarios)} usuarios:', html=html)


def _edit_user(args, user):
    identificador = args.get('email_o_username', '').strip()
    if not identificador:
        return _result('Se requiere el email o username del usuario a editar.')

    target = User.query.filter(
        (User.email == identificador) | (User.username == identificador)
    ).first()

    if not target:
        return _result(f'No se encontró usuario con email/username "{identificador}".')

    cambios = []
    if args.get('nombre'):
        target.nombre = args['nombre'].strip()
        cambios.append(f'nombre → {target.nombre}')
    if args.get('apellido'):
        target.apellido = args['apellido'].strip()
        cambios.append(f'apellido → {target.apellido}')
    if args.get('telefono') is not None:
        target.telefono = args['telefono'].strip() or None
        cambios.append(f'teléfono → {target.telefono or "(vacío)"}')
    if args.get('email'):
        new_email = args['email'].strip()
        if new_email != target.email:
            existing = User.query.filter(User.email == new_email, User.id != target.id).first()
            if existing:
                return _result(f'Ya existe otro usuario con el email "{new_email}".')
            target.email = new_email
            cambios.append(f'email → {new_email}')
    if args.get('password'):
        target.set_password(args['password'].strip())
        cambios.append('contraseña → (actualizada)')
    if args.get('rol'):
        new_rol = args['rol'].strip()
        if new_rol in ('admin', 'jefe_carrera', 'profesor_completo', 'profesor_asignatura'):
            target.rol = new_rol
            target.add_role(new_rol)
            cambios.append(f'rol → {new_rol}')
    if args.get('activo') is not None:
        target.activo = args['activo']
        cambios.append(f'activo → {target.activo}')
    if args.get('carreras') is not None:
        carreras_codigos = args['carreras']
        nuevas_carreras = []
        for codigo in carreras_codigos:
            carrera = Carrera.query.filter_by(codigo=codigo.strip().upper()).first()
            if not carrera:
                return _result(f'No se encontró la carrera con código "{codigo}".')
            nuevas_carreras.append(carrera)
        target.carreras = nuevas_carreras
        cambios.append(f'carreras → {", ".join(c.codigo for c in nuevas_carreras) or "(ninguna)"}')

    if not cambios:
        return _result('No se especificaron cambios.')

    db.session.commit()
    return _result(f'Usuario {target.get_nombre_completo()} actualizado:\n' + '\n'.join(f'- {c}' for c in cambios))


def _delete_user(args, user):
    identificador = args.get('email_o_username', '').strip()
    if not identificador:
        return _result('Se requiere el email o username del usuario a desactivar.')

    target = User.query.filter(
        (User.email == identificador) | (User.username == identificador)
    ).first()

    if not target:
        return _result(f'No se encontró usuario con email/username "{identificador}".')

    if target.id == user.id:
        return _result('No puedes desactivar tu propia cuenta.')

    if not target.activo:
        return _result(f'El usuario {target.get_nombre_completo()} ya está desactivado.')

    target.activo = False
    db.session.commit()

    return _result(f'Usuario {target.get_nombre_completo()} ({target.username}) desactivado exitosamente.')


def _create_career(args, user):
    nombre = args.get('nombre', '').strip()
    codigo = args.get('codigo', '').strip().upper()

    if not nombre or not codigo:
        return _result('Se requieren nombre y código para crear una carrera.')

    if Carrera.query.filter_by(codigo=codigo).first():
        return _result(f'Ya existe una carrera con el código "{codigo}".')

    if Carrera.query.filter_by(nombre=nombre).first():
        return _result(f'Ya existe una carrera con el nombre "{nombre}".')

    carrera = Carrera(
        nombre=nombre,
        codigo=codigo,
        descripcion=args.get('descripcion', '').strip() or None,
        facultad=args.get('facultad', '').strip() or None,
        creada_por=user.id,
    )
    db.session.add(carrera)
    db.session.commit()

    return _result(
        f'Carrera creada exitosamente:\n'
        f'- Nombre: {nombre}\n'
        f'- Código: {codigo}\n'
        f'- Facultad: {carrera.facultad or "(no especificada)"}'
    )


def _list_careers(args, user):
    carreras = Carrera.query.order_by(Carrera.nombre).all()
    if not carreras:
        return _result('No hay carreras registradas.')

    rows = []
    for c in carreras:
        estado = 'Activa' if c.activa else 'Inactiva'
        num_profesores = len([u for u in User.query.filter(
            User.rol.in_(['profesor_completo', 'profesor_asignatura']),
            User.carreras.any(Carrera.id == c.id)
        ).all()])
        jefes = User.query.filter(
            User.rol == 'jefe_carrera',
            User.carreras.any(Carrera.id == c.id)
        ).all()
        jefe_nombre = jefes[0].get_nombre_completo() if jefes else '-'
        rows.append(
            f'<tr><td>{c.codigo}</td><td>{c.nombre}</td>'
            f'<td>{c.facultad or "-"}</td><td>{estado}</td>'
            f'<td>{num_profesores}</td><td>{jefe_nombre}</td></tr>'
        )

    html = f'''<table class="table table-sm table-striped">
<thead><tr><th>Código</th><th>Nombre</th><th>Facultad</th><th>Estado</th><th># Profesores</th><th>Jefe</th></tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>'''

    return _result(f'Se encontraron {len(carreras)} carreras:', html=html)


def _create_subject(args, user):
    nombre = args.get('nombre', '').strip()
    codigo = args.get('codigo', '').strip().upper()
    cuatrimestre = args.get('cuatrimestre')
    carrera_codigo = args.get('carrera_codigo', '').strip().upper()

    if not all([nombre, codigo, cuatrimestre, carrera_codigo]):
        return _result('Se requieren nombre, código, cuatrimestre y carrera_codigo.')

    try:
        cuatrimestre = int(cuatrimestre)
    except (ValueError, TypeError):
        return _result('El cuatrimestre debe ser un número entero.')

    carrera = Carrera.query.filter_by(codigo=carrera_codigo).first()
    if not carrera:
        return _result(f'No se encontró la carrera con código "{carrera_codigo}".')

    if Materia.query.filter_by(codigo=codigo).first():
        return _result(f'Ya existe una materia con el código "{codigo}".')

    creditos = args.get('creditos', 3)
    horas_semanales = args.get('horas_semanales', 0)

    try:
        creditos = int(creditos)
        horas_semanales = int(horas_semanales)
    except (ValueError, TypeError):
        return _result('Créditos y horas semanales deben ser números enteros.')

    materia = Materia(
        nombre=nombre,
        codigo=codigo,
        cuatrimestre=cuatrimestre,
        carrera_id=carrera.id,
        creditos=creditos,
        horas_semanales=horas_semanales,
        creado_por=user.id,
    )
    db.session.add(materia)
    db.session.commit()

    return _result(
        f'Materia creada exitosamente:\n'
        f'- Nombre: {nombre}\n'
        f'- Código: {codigo}\n'
        f'- Cuatrimestre: {cuatrimestre}\n'
        f'- Carrera: {carrera.nombre} ({carrera_codigo})\n'
        f'- Créditos: {creditos}\n'
        f'- Horas semanales: {horas_semanales}'
    )


# ---- Helper: find single professor ----

def _find_single_professor(nombre):
    """Find a single professor by partial name. Returns User or _result dict on error."""
    profesores = User.query.filter(
        (User.nombre.ilike(f'%{nombre}%')) | (User.apellido.ilike(f'%{nombre}%')) |
        db.func.concat(User.nombre, ' ', User.apellido).ilike(f'%{nombre}%')
    ).filter(User.rol.in_(['profesor_completo', 'profesor_asignatura'])).all()

    if not profesores:
        return _result(f'No se encontró ningún profesor con el nombre "{nombre}".')

    if len(profesores) > 1:
        lista = '\n'.join([f"- {p.get_nombre_completo()}" for p in profesores[:10]])
        return _result(f'Se encontraron varios profesores. Por favor sé más específico:\n{lista}')

    return profesores[0]


# ---- HTML table builders ----

def _build_schedule_table_group(horarios, grupo_codigo):
    """Build an HTML schedule grid for a group (includes ALL system time slots)."""
    dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles',
                'jueves': 'Jueves', 'viernes': 'Viernes'}

    # Pre-populate grid with ALL system time slots (not just those with classes)
    all_slots = Horario.query.filter_by(activo=True).order_by(Horario.hora_inicio).all()
    grid = {}
    slot_order = []
    for slot_h in all_slots:
        slot_label = f'{slot_h.get_hora_inicio_str()} - {slot_h.get_hora_fin_str()}'
        if slot_label not in grid:
            grid[slot_label] = {d: '' for d in dias}
            slot_order.append(slot_label)

    # Fill in classes
    for h in horarios:
        slot = f'{h.get_hora_inicio_str()} - {h.get_hora_fin_str()}'
        dia = dias_map.get(h.dia_semana.lower(), h.dia_semana)
        if slot not in grid:
            grid[slot] = {d: '' for d in dias}
            slot_order.append(slot)
        if h.materia and h.profesor:
            grid[slot][dia] = f'{h.materia.nombre}<br><small>{h.profesor.get_nombre_completo()}</small>'

    rows = []
    for slot in slot_order:
        cells = ''.join([f'<td>{grid[slot].get(d, "")}</td>' for d in dias])
        rows.append(f'<tr><td><strong>{slot}</strong></td>{cells}</tr>')

    headers = ''.join([f'<th>{d}</th>' for d in dias])
    return f'''<table class="table table-sm table-bordered">
<thead class="table-dark"><tr><th>Hora</th>{headers}</tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>'''


def _build_schedule_table_professor(horarios, profesor):
    """Build an HTML schedule grid for a professor (includes ALL system time slots)."""
    dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles',
                'jueves': 'Jueves', 'viernes': 'Viernes'}

    # Pre-populate grid with ALL system time slots (not just those with classes)
    all_slots = Horario.query.filter_by(activo=True).order_by(Horario.hora_inicio).all()
    grid = {}
    slot_order = []
    for slot_h in all_slots:
        slot_label = f'{slot_h.get_hora_inicio_str()} - {slot_h.get_hora_fin_str()}'
        if slot_label not in grid:
            grid[slot_label] = {d: '' for d in dias}
            slot_order.append(slot_label)

    # Fill in classes
    for h in horarios:
        slot = f'{h.get_hora_inicio_str()} - {h.get_hora_fin_str()}'
        dia = dias_map.get(h.dia_semana.lower(), h.dia_semana)
        if slot not in grid:
            grid[slot] = {d: '' for d in dias}
            slot_order.append(slot)
        if h.materia:
            grid[slot][dia] = f'{h.materia.nombre}<br><small>Grupo: {h.grupo}</small>'

    rows = []
    for slot in slot_order:
        cells = ''.join([f'<td>{grid[slot].get(d, "")}</td>' for d in dias])
        rows.append(f'<tr><td><strong>{slot}</strong></td>{cells}</tr>')

    headers = ''.join([f'<th>{d}</th>' for d in dias])
    return f'''<table class="table table-sm table-bordered">
<thead class="table-dark"><tr><th>Hora</th>{headers}</tr></thead>
<tbody>{"".join(rows)}</tbody>
</table>'''


# ---- Excel generation helpers ----

def _generate_group_excel(horarios, grupo_codigo, user):
    """Generate an Excel file for a group schedule and return download URL."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = f'Horario {grupo_codigo}'

        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Horario del Grupo {grupo_codigo}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Column headers
        dias = ['Hora', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        for col, dia in enumerate(dias, 1):
            cell = ws.cell(row=3, column=col, value=dia)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='0070C0')
            cell.alignment = Alignment(horizontal='center')

        # Build grid
        dias_map = {'lunes': 1, 'martes': 2, 'miercoles': 3, 'jueves': 4, 'viernes': 5}
        grid = {}
        for h in horarios:
            slot = f'{h.get_hora_inicio_str()} - {h.get_hora_fin_str()}'
            if slot not in grid:
                grid[slot] = [''] * 5
            col_idx = dias_map.get(h.dia_semana.lower())
            if col_idx is not None and h.materia:
                prof_name = h.profesor.get_nombre_completo() if h.profesor else ''
                grid[slot][col_idx - 1] = f'{h.materia.nombre}\n{prof_name}'

        for row_idx, slot in enumerate(sorted(grid.keys()), 4):
            ws.cell(row=row_idx, column=1, value=slot).font = Font(bold=True)
            for col_idx, val in enumerate(grid[slot], 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Adjust widths
        ws.column_dimensions['A'].width = 18
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 22

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        file_id = file_manager.save_file(
            buf.getvalue(),
            f'horario_{grupo_codigo}.xlsx',
            user.id,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        return f'/api/chatbot/download/{file_id}'
    except Exception as e:
        logger.error(f"Error generando Excel de grupo: {e}")
        return None


def _generate_professor_excel(profesor, user):
    """Generate an Excel file for a professor schedule using FDA format and return download URL."""
    try:
        from app_pkg.helpers.export_helpers import generar_excel_profesor_buffer

        buf, filename = generar_excel_profesor_buffer(profesor.get_nombre_completo())
        if buf is None:
            return None

        file_id = file_manager.save_file(
            buf.getvalue(),
            filename,
            user.id,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        return f'/api/chatbot/download/{file_id}'
    except Exception as e:
        logger.error(f"Error generando Excel de profesor: {e}")
        return None


def _generate_professor_downloads(profesor, user):
    """Generate both Excel and PDF for a professor schedule. Returns list of {url, label}."""
    downloads = []

    # Excel
    excel_url = _generate_professor_excel(profesor, user)
    if excel_url:
        downloads.append({'url': excel_url, 'label': 'Excel'})

    # PDF
    try:
        from app_pkg.helpers.export_helpers import generar_pdf_profesor_buffer

        buf, filename = generar_pdf_profesor_buffer(profesor.get_nombre_completo())
        if buf is not None:
            file_id = file_manager.save_file(
                buf.getvalue(),
                filename,
                user.id,
                'application/pdf'
            )
            downloads.append({'url': f'/api/chatbot/download/{file_id}', 'label': 'PDF'})
    except Exception as e:
        logger.error(f"Error generando PDF de profesor: {e}")

    return downloads if downloads else None


# ---- CSV generation helpers ----

def _generate_group_csv(horarios, grupo_codigo, user):
    """Generate a CSV file for a group schedule and return download URL."""
    try:
        dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles',
                    'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Día', 'Hora Inicio', 'Hora Fin', 'Materia', 'Profesor'])

        for h in sorted(horarios, key=lambda x: (x.dia_semana, x.get_hora_inicio_str())):
            dia = dias_map.get(h.dia_semana.lower(), h.dia_semana)
            materia = h.materia.nombre if h.materia else ''
            profesor = h.profesor.get_nombre_completo() if h.profesor else ''
            writer.writerow([dia, h.get_hora_inicio_str(), h.get_hora_fin_str(), materia, profesor])

        csv_bytes = output.getvalue().encode('utf-8-sig')
        file_id = file_manager.save_file(
            csv_bytes,
            f'horario_{grupo_codigo}.csv',
            user.id,
            'text/csv'
        )
        return f'/api/chatbot/download/{file_id}'
    except Exception as e:
        logger.error(f"Error generando CSV de grupo: {e}")
        return None


def _generate_professor_csv(profesor, user):
    """Generate a CSV file for a professor schedule and return download URL."""
    try:
        nombre_completo = profesor.get_nombre_completo()
        horarios = HorarioAcademico.query.filter_by(
            profesor_id=profesor.id, activo=True
        ).all()

        if not horarios:
            return None

        dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles',
                    'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado'}

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Día', 'Hora Inicio', 'Hora Fin', 'Materia', 'Grupo'])

        for h in sorted(horarios, key=lambda x: (x.dia_semana, x.get_hora_inicio_str())):
            dia = dias_map.get(h.dia_semana.lower(), h.dia_semana)
            materia = h.materia.nombre if h.materia else ''
            writer.writerow([dia, h.get_hora_inicio_str(), h.get_hora_fin_str(), materia, h.grupo or ''])

        csv_bytes = output.getvalue().encode('utf-8-sig')
        file_id = file_manager.save_file(
            csv_bytes,
            f'horario_{nombre_completo.replace(" ", "_")}.csv',
            user.id,
            'text/csv'
        )
        return f'/api/chatbot/download/{file_id}'
    except Exception as e:
        logger.error(f"Error generando CSV de profesor: {e}")
        return None
