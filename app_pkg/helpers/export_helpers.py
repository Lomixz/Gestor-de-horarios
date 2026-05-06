"""
Export helper functions.

Contains functions for generating Excel/PDF exports in UPTEX FDA format.
Extracted from app.py lines 46-69, 571-1167, 8780-9213.
"""
import os
import logging
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XlImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RlImage, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from PIL import Image as PILImage

from .schedule_helpers import obtener_periodo_actual

logger = logging.getLogger('sistema_academico')


def convertir_imagen_para_excel(ruta_imagen):
    """
    Convierte imágenes en formatos no soportados por openpyxl (como .webp)
    a PNG en memoria y devuelve un XlImage compatible.
    Retorna None si no se puede procesar.
    """
    try:
        ext = os.path.splitext(ruta_imagen)[1].lower()
        if ext in ('.webp', '.bmp', '.tiff', '.tif'):
            # Convertir a PNG en memoria
            pil_img = PILImage.open(ruta_imagen)
            if pil_img.mode in ('RGBA', 'LA', 'P'):
                pil_img = pil_img.convert('RGBA')
            else:
                pil_img = pil_img.convert('RGB')
            img_buffer = BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            return XlImage(img_buffer)
        else:
            return XlImage(ruta_imagen)
    except Exception as e:
        logger.error(f"Error al convertir imagen: {e}")
        return None


def generar_excel_formato_fda(datos_profesor, periodo=None, año=None):
    """
    Genera Excel con formato EXACTO de la plantilla HORARIO ACTUAL.xlsx de UPTEX.
    Replica fielmente: logo, colores verdes #00B050, estructura de celdas combinadas,
    tabla de horarios con 2 filas por hora, sección de firmas.

    Args:
        datos_profesor: Diccionario con info del profesor y sus clases
        periodo: Periodo académico (opcional, se calcula automáticamente)
        año: Año del plan (opcional, se calcula automáticamente)

    Returns:
        BytesIO: Buffer con el archivo Excel generado
    """
    from models import ConfiguracionSistema, User

    # Obtener periodo automáticamente si no se especifica
    if periodo is None or año is None:
        periodo, año = obtener_periodo_actual()

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # ========== COLOR VERDE EXACTO DE LA PLANTILLA (#00B050) ==========
    verde_uptex = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

    # ========== ESTILOS ==========
    titulo_font = Font(bold=True, name='Century Gothic', size=14, color="FFFFFF")
    header_font = Font(bold=True, name='Century Gothic', size=10)
    normal_font = Font(name='Century Gothic', size=10)
    small_font = Font(name='Century Gothic', size=9)
    tiny_font = Font(name='Century Gothic', size=8)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== ANCHOS DE COLUMNA (EXACTOS DE LA PLANTILLA) ==========
    ws.column_dimensions['A'].width = 9.5
    ws.column_dimensions['B'].width = 22.0
    ws.column_dimensions['C'].width = 22.0
    ws.column_dimensions['D'].width = 22.0
    ws.column_dimensions['E'].width = 22.0
    ws.column_dimensions['F'].width = 15.5
    ws.column_dimensions['G'].width = 0.5  # Columna casi oculta
    ws.column_dimensions['H'].width = 7
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['J'].width = 17
    ws.column_dimensions['K'].width = 13
    ws.column_dimensions['L'].width = 13

    # ========== 1. LOGO UPTEX (A1) ==========
    logo_path = os.path.join('static', 'images', 'logo.png')
    if os.path.exists(logo_path):
        try:
            img = convertir_imagen_para_excel(logo_path)
            if img:
                img.width = 150
                img.height = 55
                ws.add_image(img, 'A1')
        except Exception as e:
            logger.error(f"Error al cargar logo: {e}")

    # ========== 2. FILA 1: "Carga Horaria" (B1:L1 - verde) ==========
    ws.merge_cells('B1:L1')
    ws['B1'] = "Carga Horaria"
    ws['B1'].font = titulo_font
    ws['B1'].alignment = center_align
    ws['B1'].fill = verde_uptex
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws[f'{col}1'].fill = verde_uptex

    # ========== 3. FILA 2: Área, Vigencia, Código (verde con texto blanco) ==========
    label_verde_font = Font(bold=True, name='Century Gothic', size=10, color="FFFFFF")

    ws.merge_cells('B2:C2')
    ws['B2'] = "Área: Dirección Academica"
    ws['B2'].font = label_verde_font
    ws['B2'].alignment = center_align
    ws['B2'].fill = verde_uptex
    ws['C2'].fill = verde_uptex

    ws.merge_cells('D2:E2')
    ws['D2'] = "Vigencia: "
    ws['D2'].font = label_verde_font
    ws['D2'].alignment = center_align
    ws['D2'].fill = verde_uptex
    ws['E2'].fill = verde_uptex

    ws.merge_cells('F2:L2')
    ws['F2'] = "Código: FDA-02.5"
    ws['F2'].font = label_verde_font
    ws['F2'].alignment = center_align
    ws['F2'].fill = verde_uptex
    for col in ['G', 'H', 'I', 'J', 'K', 'L']:
        ws[f'{col}2'].fill = verde_uptex

    # ========== 4. FILA 4: Nombre, Prof. Asignatura, Prof. TC ==========
    ws['A4'] = "Nombre:"
    ws['A4'].font = header_font

    ws.merge_cells('B4:E4')
    ws['B4'] = datos_profesor['info']['nombre_completo']
    ws['B4'].font = normal_font

    ws['F4'] = "Prof. Asignatura"
    ws['F4'].font = label_verde_font
    ws['F4'].fill = verde_uptex
    ws['F4'].alignment = center_align

    ws['H4'] = "x" if not datos_profesor['info']['es_tc'] else ""
    ws['H4'].font = normal_font
    ws['H4'].alignment = center_align
    ws['H4'].border = thin_border

    ws.merge_cells('J4:K4')
    ws['J4'] = "Prof. Tiempo Completo"
    ws['J4'].font = label_verde_font
    ws['J4'].fill = verde_uptex
    ws['J4'].alignment = center_align
    ws['K4'].fill = verde_uptex

    ws['L4'] = "X" if datos_profesor['info']['es_tc'] else ""
    ws['L4'].font = normal_font
    ws['L4'].alignment = center_align
    ws['L4'].border = thin_border

    # ========== 5. FILA 6: Periodo, Fecha de Inicio, Plan de Estudios ==========
    ws['A6'] = "Periodo:"
    ws['A6'].font = header_font

    ws['B6'] = periodo
    ws['B6'].font = normal_font
    ws['B6'].alignment = center_align

    ws['C6'] = "Fecha de Inicio:"
    ws['C6'].font = header_font

    # Obtener fecha de inicio desde configuración del sistema
    config_fecha_inicio = ConfiguracionSistema.query.filter_by(clave='fecha_inicio_periodo').first()
    fecha_inicio_valor = config_fecha_inicio.valor if config_fecha_inicio and config_fecha_inicio.valor else ""

    ws.merge_cells('D6:E6')
    ws['D6'] = fecha_inicio_valor
    ws['D6'].font = normal_font
    ws['D6'].alignment = center_align

    ws['F6'] = "Plan de Estudios:"
    ws['F6'].font = header_font

    ws.merge_cells('G6:L6')
    ws['G6'] = año
    ws['G6'].font = normal_font
    ws['G6'].alignment = center_align

    # ========== 6. FILA 8-9: Instrucciones ==========
    ws.merge_cells('A8:L9')
    ws['A8'] = "Instrucciones: Introducir nombre de la Asignatura, Salón y Grupo dentro de la celda correspondiente al día y la hora que será impartida."
    ws['A8'].font = Font(bold=True, name='Century Gothic', size=8)
    ws['A8'].alignment = left_align
    ws['A8'].alignment = left_align

    # ========== 7. FILA 11: ENCABEZADOS DE TABLA (verde) ==========
    # Horario
    ws['A11'] = "Horario "
    ws['A11'].font = header_font
    ws['A11'].alignment = center_align
    ws['A11'].fill = verde_uptex
    ws['A11'].border = thin_border

    # Lunes a Jueves
    for col, dia in [('B', 'Lunes '), ('C', 'Martes'), ('D', 'Miercoles'), ('E', 'Jueves ')]:
        ws[f'{col}11'] = dia
        ws[f'{col}11'].font = header_font
        ws[f'{col}11'].alignment = center_align
        ws[f'{col}11'].fill = verde_uptex
        ws[f'{col}11'].border = thin_border

    # Viernes (F11:H11)
    ws.merge_cells('F11:H11')
    ws['F11'] = "Viernes "
    ws['F11'].font = header_font
    ws['F11'].alignment = center_align
    ws['F11'].fill = verde_uptex
    ws['F11'].border = thin_border
    for col in ['G', 'H']:
        ws[f'{col}11'].fill = verde_uptex
        ws[f'{col}11'].border = thin_border

    # Sábado (I11:L11)
    ws.merge_cells('I11:L11')
    ws['I11'] = "Sábado "
    ws['I11'].font = header_font
    ws['I11'].alignment = center_align
    ws['I11'].fill = verde_uptex
    ws['I11'].border = thin_border
    for col in ['J', 'K', 'L']:
        ws[f'{col}11'].fill = verde_uptex
        ws[f'{col}11'].border = thin_border

    # ========== 8. FILAS DE HORARIOS (2 filas por hora) ==========
    # Use actual time slots from the system instead of hardcoded hours
    from models import Horario as HorarioModel
    horarios_sistema = HorarioModel.query.filter_by(activo=True).order_by(HorarioModel.hora_inicio).all()
    horas_set = []
    seen = set()
    for hs in horarios_sistema:
        h_str = f"{hs.hora_inicio.hour:02d}:00"
        if h_str not in seen:
            seen.add(h_str)
            horas_set.append((h_str, hs.hora_inicio, hs.hora_fin))
    # Build simple hour strings for matching, but keep full range for display
    horas = [h[0] for h in horas_set]

    # Preparar datos de clases por día y hora
    clases_por_dia_hora = {}
    for clase in datos_profesor['clases']:
        dia_raw = clase['dia_raw'].lower()
        hora_inicio = clase['hora_inicio']

        if dia_raw == 'miércoles':
            dia_raw = 'miercoles'
        if dia_raw == 'sábado':
            dia_raw = 'sabado'

        try:
            h_ini = int(hora_inicio.split(':')[0])
            hora_fin = clase['hora_fin']
            h_fin = int(hora_fin.split(':')[0])
            for h in range(h_ini, h_fin):
                hora_key = f"{h:02d}:00"
                key = (dia_raw, hora_key)
                if key not in clases_por_dia_hora:
                    clases_por_dia_hora[key] = clase
        except Exception:
            pass

    # Add PTC activities for TC professors
    from models import ActividadPTC as ActPTCModel
    ptc_por_dia_hora = {}
    profesor_id = datos_profesor['info'].get('id')
    es_tc = datos_profesor['info'].get('es_tc', False)
    if profesor_id and es_tc:
        ptc_activities = ActPTCModel.query.filter_by(profesor_id=profesor_id, activo=True).all()
        ptc_label_map = {'asesoria': 'ASESORÍA', 'tutoria': 'TUTORÍA', 'gestion': 'GESTIÓN'}
        for p in ptc_activities:
            if p.horario and p.horario.hora_inicio:
                dia_raw = p.dia_semana.lower()
                if dia_raw == 'miércoles':
                    dia_raw = 'miercoles'
                if dia_raw == 'sábado':
                    dia_raw = 'sabado'
                hora_key = f"{p.horario.hora_inicio.hour:02d}:00"
                key = (dia_raw, hora_key)
                ptc_por_dia_hora[key] = {
                    'tipo': p.tipo_actividad,
                    'label': ptc_label_map.get(p.tipo_actividad, p.tipo_actividad.upper()),
                    'notas': p.notas or ''
                }

    dia_col_map = {
        'lunes': 'B', 'martes': 'C', 'miercoles': 'D',
        'jueves': 'E', 'viernes': 'F', 'sabado': 'I'
    }

    # PTC fill styles for Excel
    ptc_fill_map = {
        'asesoria': PatternFill(start_color="28A745", end_color="28A745", fill_type="solid"),
        'tutoria': PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
        'gestion': PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid"),
    }
    ptc_font_white = Font(bold=True, name='Century Gothic', size=7, color="FFFFFF")
    ptc_font_dark = Font(bold=True, name='Century Gothic', size=7, color="212529")

    def _write_cell(ws, col, fila_inicio, fila_fin, dia, hora, merge_end_col=None):
        """Write class or PTC content into a cell range."""
        key = (dia, hora)
        if merge_end_col:
            ws.merge_cells(f'{col}{fila_inicio}:{merge_end_col}{fila_fin}')
        else:
            ws.merge_cells(f'{col}{fila_inicio}:{col}{fila_fin}')
        
        if key in clases_por_dia_hora:
            clase = clases_por_dia_hora[key]
            ws[f'{col}{fila_inicio}'] = f"{clase['asignatura']} {clase['grupo']}"
            ws[f'{col}{fila_inicio}'].font = tiny_font
        elif key in ptc_por_dia_hora:
            ptc = ptc_por_dia_hora[key]
            ws[f'{col}{fila_inicio}'] = ptc['label']
            ws[f'{col}{fila_inicio}'].fill = ptc_fill_map.get(ptc['tipo'], PatternFill())
            ws[f'{col}{fila_inicio}'].font = ptc_font_dark if ptc['tipo'] == 'tutoria' else ptc_font_white
            # Also fill merged cells
            if merge_end_col:
                for c in range(ord(col), ord(merge_end_col) + 1):
                    for r in [fila_inicio, fila_fin]:
                        ws[f'{chr(c)}{r}'].fill = ptc_fill_map.get(ptc['tipo'], PatternFill())
            else:
                ws[f'{col}{fila_fin}'].fill = ptc_fill_map.get(ptc['tipo'], PatternFill())
        else:
            ws[f'{col}{fila_inicio}'].font = tiny_font
        
        ws[f'{col}{fila_inicio}'].alignment = center_align
        ws[f'{col}{fila_inicio}'].border = thin_border
        ws[f'{col}{fila_fin}'].border = thin_border

    fila_actual = 12
    for hora in horas:
        fila_inicio = fila_actual
        fila_fin = fila_actual + 1

        # Hora (A - combinar 2 filas)
        ws.merge_cells(f'A{fila_inicio}:A{fila_fin}')
        
        # Formatear como rango de horas (ej: 07:00 - 08:00)
        h_int = int(hora.split(':')[0])
        hora_rango = f"{h_int:02d}:00 - {h_int+1:02d}:00"
        
        ws[f'A{fila_inicio}'] = hora_rango
        ws[f'A{fila_inicio}'].font = tiny_font
        ws[f'A{fila_inicio}'].alignment = center_align
        ws[f'A{fila_inicio}'].border = thin_border
        ws[f'A{fila_fin}'].border = thin_border

        # Lunes a Jueves (combinar 2 filas cada uno)
        for dia, col in [('lunes', 'B'), ('martes', 'C'), ('miercoles', 'D'), ('jueves', 'E')]:
            _write_cell(ws, col, fila_inicio, fila_fin, dia, hora)

        # Viernes (F:H - combinar)
        _write_cell(ws, 'F', fila_inicio, fila_fin, 'viernes', hora, merge_end_col='H')
        for col in ['G', 'H']:
            ws[f'{col}{fila_inicio}'].border = thin_border
            ws[f'{col}{fila_fin}'].border = thin_border

        # Sábado (I:L - combinar)
        _write_cell(ws, 'I', fila_inicio, fila_fin, 'sabado', hora, merge_end_col='L')
        for col in ['J', 'K', 'L']:
            ws[f'{col}{fila_inicio}'].border = thin_border
            ws[f'{col}{fila_fin}'].border = thin_border

        fila_actual += 2

    # ========== 9. ALTURA DE FILAS ==========
    for row in range(12, fila_actual):
        ws.row_dimensions[row].height = 20

    # ========== 10. TABLA DE TIPO DE HORAS ==========
    # Calcular total de horas
    # Calcular horas de impartición de curso (suma real de horas)
    horas_imparticion = sum(c['horas_totales'] for c in datos_profesor['clases'])
    horas_imparticion = int(horas_imparticion) if horas_imparticion == int(horas_imparticion) else horas_imparticion

    # Obtener horas adicionales de TC desde configuración (solo para profesores TC)
    profesor_id = datos_profesor['info'].get('id')
    es_tc = datos_profesor['info'].get('es_tc', False)

    horas_asesoria = 0
    horas_tutoria = 0
    horas_gestion = 0
    horas_dual = 0
    horas_investigacion = 0

    if profesor_id and es_tc and ptc_por_dia_hora:
        # Count from actual PTC activities
        horas_asesoria = sum(1 for v in ptc_por_dia_hora.values() if v['tipo'] == 'asesoria')
        horas_tutoria = sum(1 for v in ptc_por_dia_hora.values() if v['tipo'] == 'tutoria')
        horas_gestion = sum(1 for v in ptc_por_dia_hora.values() if v['tipo'] == 'gestion')
        horas_dual = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor_id}_dual', 0) or 0)
        horas_investigacion = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor_id}_investigacion', 0) or 0)
    elif profesor_id:
        horas_asesoria = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor_id}_asesoria', 0) or 0)
        horas_tutoria = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor_id}_tutoria', 0) or 0)
        horas_gestion = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor_id}_gestion', 0) or 0)
        horas_dual = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor_id}_dual', 0) or 0)
        horas_investigacion = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor_id}_investigacion', 0) or 0)

    total_horas = horas_imparticion + horas_asesoria + horas_tutoria + horas_gestion + horas_dual + horas_investigacion

    fila_tabla = fila_actual + 1

    # Header de la tabla (verde)
    ws.merge_cells(f'A{fila_tabla}:B{fila_tabla}')
    ws[f'A{fila_tabla}'] = "Tipo de Horas"
    ws[f'A{fila_tabla}'].font = Font(bold=True, name='Century Gothic', size=10, color="FFFFFF")
    ws[f'A{fila_tabla}'].alignment = center_align
    ws[f'A{fila_tabla}'].fill = verde_uptex
    ws[f'A{fila_tabla}'].border = thin_border
    ws[f'B{fila_tabla}'].fill = verde_uptex
    ws[f'B{fila_tabla}'].border = thin_border

    ws[f'C{fila_tabla}'] = "Horas"
    ws[f'C{fila_tabla}'].font = Font(bold=True, name='Century Gothic', size=10, color="FFFFFF")
    ws[f'C{fila_tabla}'].alignment = center_align
    ws[f'C{fila_tabla}'].fill = verde_uptex
    ws[f'C{fila_tabla}'].border = thin_border

    # Filas de tipos de horas
    tipos_horas = [
        ('Impartición de Curso', horas_imparticion),
        ('Asesoría', horas_asesoria),
        ('Tutoría', horas_tutoria),
        ('Apoyo a la Gestión', horas_gestion),
        ('Dual', horas_dual),
        ('Investigación', horas_investigacion),
    ]

    for i, (tipo, horas_val) in enumerate(tipos_horas, 1):
        fila_tipo = fila_tabla + i
        ws.merge_cells(f'A{fila_tipo}:B{fila_tipo}')
        ws[f'A{fila_tipo}'] = tipo
        ws[f'A{fila_tipo}'].font = normal_font
        ws[f'A{fila_tipo}'].alignment = left_align
        ws[f'A{fila_tipo}'].border = thin_border
        ws[f'B{fila_tipo}'].border = thin_border
        # Mostrar 0 como vacío para que se vea más limpio
        ws[f'C{fila_tipo}'] = horas_val if horas_val != 0 else ''
        ws[f'C{fila_tipo}'].font = normal_font
        ws[f'C{fila_tipo}'].alignment = center_align
        ws[f'C{fila_tipo}'].border = thin_border

    # Total de horas
    fila_total = fila_tabla + len(tipos_horas) + 1
    ws.merge_cells(f'A{fila_total}:B{fila_total}')
    ws[f'A{fila_total}'] = "Total de Horas"
    ws[f'A{fila_total}'].font = header_font
    ws[f'A{fila_total}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'A{fila_total}'].border = thin_border
    ws[f'B{fila_total}'].border = thin_border
    ws[f'C{fila_total}'] = total_horas
    ws[f'C{fila_total}'].font = header_font
    ws[f'C{fila_total}'].alignment = center_align
    ws[f'C{fila_total}'].border = thin_border

    # Nota
    fila_nota = fila_total + 1
    ws.merge_cells(f'A{fila_nota}:E{fila_nota}')
    ws[f'A{fila_nota}'] = "*Solo llenar en caso de ser Profesor de Tiempo Completo"
    ws[f'A{fila_nota}'].font = Font(italic=True, name='Century Gothic', size=7)

    # ========== 11. SECCIÓN DE FIRMAS ==========
    gris_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    firma_header_font = Font(bold=True, name='Century Gothic', size=10)

    fila_firma = fila_nota + 2

    config_director = ConfiguracionSistema.query.filter_by(clave='director_academico_nombre').first()
    config_responsable = ConfiguracionSistema.query.filter_by(clave='responsable_pa_nombre').first()

    nombre_director = config_director.valor if config_director and config_director.valor else ""
    nombre_responsable = config_responsable.valor if config_responsable and config_responsable.valor else ""

    # Encabezados de firma con fondo gris
    # Elaboró
    ws.merge_cells(f'A{fila_firma}:C{fila_firma}')
    ws[f'A{fila_firma}'] = "Elaboró:"
    ws[f'A{fila_firma}'].font = firma_header_font
    ws[f'A{fila_firma}'].alignment = center_align
    ws[f'A{fila_firma}'].fill = gris_fill
    ws[f'A{fila_firma}'].border = thin_border
    for col in ['B', 'C']:
        ws[f'{col}{fila_firma}'].fill = gris_fill
        ws[f'{col}{fila_firma}'].border = thin_border

    # Autorizó
    ws.merge_cells(f'D{fila_firma}:F{fila_firma}')
    ws[f'D{fila_firma}'] = "Autorizó:"
    ws[f'D{fila_firma}'].font = firma_header_font
    ws[f'D{fila_firma}'].alignment = center_align
    ws[f'D{fila_firma}'].fill = gris_fill
    ws[f'D{fila_firma}'].border = thin_border
    for col in ['E', 'F']:
        ws[f'{col}{fila_firma}'].fill = gris_fill
        ws[f'{col}{fila_firma}'].border = thin_border

    # Recibió
    ws.merge_cells(f'G{fila_firma}:L{fila_firma}')
    ws[f'G{fila_firma}'] = "Recibió:"
    ws[f'G{fila_firma}'].font = firma_header_font
    ws[f'G{fila_firma}'].alignment = center_align
    ws[f'G{fila_firma}'].fill = gris_fill
    ws[f'G{fila_firma}'].border = thin_border
    for col in ['H', 'I', 'J', 'K', 'L']:
        ws[f'{col}{fila_firma}'].fill = gris_fill
        ws[f'{col}{fila_firma}'].border = thin_border

    # Nombres
    fila_nombres = fila_firma + 1
    ws.merge_cells(f'A{fila_nombres}:C{fila_nombres}')
    ws[f'A{fila_nombres}'] = datos_profesor['info']['nombre_completo']
    ws[f'A{fila_nombres}'].font = normal_font
    ws[f'A{fila_nombres}'].alignment = center_align
    ws[f'A{fila_nombres}'].border = thin_border
    for col in ['B', 'C']:
        ws[f'{col}{fila_nombres}'].border = thin_border

    ws.merge_cells(f'D{fila_nombres}:F{fila_nombres}')
    ws[f'D{fila_nombres}'] = nombre_director
    ws[f'D{fila_nombres}'].font = normal_font
    ws[f'D{fila_nombres}'].alignment = center_align
    ws[f'D{fila_nombres}'].border = thin_border
    for col in ['E', 'F']:
        ws[f'{col}{fila_nombres}'].border = thin_border

    ws.merge_cells(f'G{fila_nombres}:L{fila_nombres}')
    ws[f'G{fila_nombres}'] = nombre_responsable
    ws[f'G{fila_nombres}'].font = normal_font
    ws[f'G{fila_nombres}'].alignment = center_align
    ws[f'G{fila_nombres}'].border = thin_border
    for col in ['H', 'I', 'J', 'K', 'L']:
        ws[f'{col}{fila_nombres}'].border = thin_border

    # Cargos
    fila_cargos = fila_nombres + 1
    cargo_profesor = "PROFESOR DE TIEMPO COMPLETO" if datos_profesor['info']['es_tc'] else "PROFESOR DE ASIGNATURA"

    ws.merge_cells(f'A{fila_cargos}:C{fila_cargos}')
    ws[f'A{fila_cargos}'] = cargo_profesor
    ws[f'A{fila_cargos}'].font = normal_font
    ws[f'A{fila_cargos}'].alignment = center_align
    ws[f'A{fila_cargos}'].border = thin_border
    for col in ['B', 'C']:
        ws[f'{col}{fila_cargos}'].border = thin_border

    ws.merge_cells(f'D{fila_cargos}:F{fila_cargos}')
    ws[f'D{fila_cargos}'] = "Director Académico"
    ws[f'D{fila_cargos}'].font = normal_font
    ws[f'D{fila_cargos}'].alignment = center_align
    ws[f'D{fila_cargos}'].border = thin_border
    for col in ['E', 'F']:
        ws[f'{col}{fila_cargos}'].border = thin_border

    ws.merge_cells(f'G{fila_cargos}:L{fila_cargos}')
    ws[f'G{fila_cargos}'] = "Responsable del PA"
    ws[f'G{fila_cargos}'].font = normal_font
    ws[f'G{fila_cargos}'].alignment = center_align
    ws[f'G{fila_cargos}'].border = thin_border
    for col in ['H', 'I', 'J', 'K', 'L']:
        ws[f'{col}{fila_cargos}'].border = thin_border

    # Espacio para firma (fila vacía con bordes)
    fila_espacio = fila_cargos + 1
    ws.row_dimensions[fila_espacio].height = 45
    ws.merge_cells(f'A{fila_espacio}:C{fila_espacio}')
    ws.merge_cells(f'D{fila_espacio}:F{fila_espacio}')
    ws.merge_cells(f'G{fila_espacio}:L{fila_espacio}')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws[f'{col}{fila_espacio}'].border = thin_border

    # Fila extra de espacio para que la firma no tape cargos
    fila_espacio2 = fila_espacio + 1
    ws.row_dimensions[fila_espacio2].height = 15
    ws.merge_cells(f'A{fila_espacio2}:C{fila_espacio2}')
    ws.merge_cells(f'D{fila_espacio2}:F{fila_espacio2}')
    ws.merge_cells(f'G{fila_espacio2}:L{fila_espacio2}')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws[f'{col}{fila_espacio2}'].border = thin_border

    # Etiquetas "Firma"
    fila_label_firma = fila_espacio2 + 1
    ws.merge_cells(f'A{fila_label_firma}:C{fila_label_firma}')
    ws[f'A{fila_label_firma}'] = "Firma"
    ws[f'A{fila_label_firma}'].font = normal_font
    ws[f'A{fila_label_firma}'].alignment = center_align
    ws[f'A{fila_label_firma}'].border = thin_border
    for col in ['B', 'C']:
        ws[f'{col}{fila_label_firma}'].border = thin_border

    ws.merge_cells(f'D{fila_label_firma}:F{fila_label_firma}')
    ws[f'D{fila_label_firma}'] = "Firma"
    ws[f'D{fila_label_firma}'].font = normal_font
    ws[f'D{fila_label_firma}'].alignment = center_align
    ws[f'D{fila_label_firma}'].border = thin_border
    for col in ['E', 'F']:
        ws[f'{col}{fila_label_firma}'].border = thin_border

    ws.merge_cells(f'G{fila_label_firma}:L{fila_label_firma}')
    ws[f'G{fila_label_firma}'] = "Firma"
    ws[f'G{fila_label_firma}'].font = normal_font
    ws[f'G{fila_label_firma}'].alignment = center_align
    ws[f'G{fila_label_firma}'].border = thin_border
    for col in ['H', 'I', 'J', 'K', 'L']:
        ws[f'{col}{fila_label_firma}'].border = thin_border

    # ========== 12. INSERTAR FIRMAS DIGITALES SI EXISTEN ==========
    fila_firma_img = fila_espacio
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker as AnchorMk
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU

    firma_w = 140   # píxeles
    firma_h = 45    # píxeles
    firma_ext = XDRPositiveSize2D(pixels_to_EMU(firma_w), pixels_to_EMU(firma_h))
    row_idx = fila_firma_img - 1  # 0-based row index
    row_off = pixels_to_EMU(3)    # pequeño margen superior

    # Anchos de columna en px (aprox 7.5 px por char width en Excel)
    # A=9.5, B=17.2, C=18.2 → total A-C ≈ 337px → center offset = (337-140)/2 ≈ 98px
    # A es ~71px, así que empezamos en col B (idx=1) con offset = 98-71 = 27px
    # D=17.3, E=16.0, F=15.5 → total D-F ≈ 366px → center offset = (366-140)/2 ≈ 113px
    # D es ~130px, así que empezamos en col D (idx=3) con offset = 113px
    # G=0.5, H=7, I=13, J=17, K=13, L=13 → total ≈ 479px → center offset = (479-140)/2 ≈ 170px
    # G(4)+H(53)+I(98)=155px, así que empezamos en col J (idx=9) con offset = 170-155 = 15px

    def crear_anchor_firma(col_idx, col_off_px):
        marker = AnchorMk(col=col_idx, colOff=pixels_to_EMU(col_off_px), row=row_idx, rowOff=row_off)
        return OneCellAnchor(_from=marker, ext=firma_ext)

    # 1. Firma del profesor (Elaboró) — centrada en A-C
    profesor = User.query.get(datos_profesor['info']['id'])
    if profesor and profesor.firma:
        firma_path = os.path.join('static', 'uploads', 'firmas', profesor.firma)
        if os.path.exists(firma_path):
            try:
                firma_img = convertir_imagen_para_excel(firma_path)
                if firma_img:
                    firma_img.anchor = crear_anchor_firma(1, 55)
                    ws.add_image(firma_img)
            except Exception as e:
                logger.error(f"Error al cargar firma del profesor: {e}")

    # 2. Firma del Director Académico (Autorizó) — centrada en D-F
    config_firma_director = ConfiguracionSistema.query.filter_by(clave='director_academico_firma').first()
    if config_firma_director and config_firma_director.valor:
        firma_director_path = os.path.join('static', 'uploads', 'firmas', config_firma_director.valor)
        if os.path.exists(firma_director_path):
            try:
                firma_dir_img = convertir_imagen_para_excel(firma_director_path)
                if firma_dir_img:
                    firma_dir_img.anchor = crear_anchor_firma(3, 113)
                    ws.add_image(firma_dir_img)
            except Exception as e:
                logger.error(f"Error al cargar firma del director: {e}")

    # 3. Firma del Responsable del PA (Recibió) — centrada en G-L
    config_firma_responsable = ConfiguracionSistema.query.filter_by(clave='responsable_pa_firma').first()
    if config_firma_responsable and config_firma_responsable.valor:
        firma_resp_path = os.path.join('static', 'uploads', 'firmas', config_firma_responsable.valor)
        if os.path.exists(firma_resp_path):
            try:
                firma_resp_img = convertir_imagen_para_excel(firma_resp_path)
                if firma_resp_img:
                    firma_resp_img.anchor = crear_anchor_firma(9, 15)
                    ws.add_image(firma_resp_img)
            except Exception as e:
                logger.error(f"Error al cargar firma del responsable: {e}")

    # ========== 13. GUARDAR EN BUFFER ==========
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_excel_profesor_buffer(profesor_nombre):
    """
    Genera un BytesIO con el Excel del horario del profesor en formato FDA.
    Retorna (BytesIO, filename) o (None, error_message).
    """
    from models import User, HorarioAcademico

    profesor = User.query.filter(
        (User.nombre + ' ' + User.apellido) == profesor_nombre
    ).first()

    if not profesor:
        return None, "Profesor no encontrado"

    asignaciones = HorarioAcademico.query.filter_by(profesor_id=profesor.id, activo=True).all()

    dias_map = {'lunes': 'lunes', 'martes': 'martes', 'miercoles': 'miercoles',
                'miércoles': 'miercoles', 'jueves': 'jueves', 'viernes': 'viernes',
                'sabado': 'sabado', 'sábado': 'sabado'}

    es_tc = profesor.is_profesor_completo()

    datos_profesor = {
        'info': {
            'id': profesor.id,
            'nombre_completo': profesor.get_nombre_completo(),
            'es_tc': es_tc
        },
        'clases': []
    }

    for a in asignaciones:
        if not all([a.materia, a.horario]):
            continue

        dia_raw = dias_map.get(a.dia_semana.lower(), a.dia_semana.lower())
        grupo_codigo = a.grupo if a.grupo else "N/A"

        duracion_horas = (a.horario.hora_fin.hour - a.horario.hora_inicio.hour) + \
                         (a.horario.hora_fin.minute - a.horario.hora_inicio.minute) / 60.0

        detalle_clase = {
            'clave': a.materia.codigo if hasattr(a.materia, 'codigo') else '',
            'asignatura': a.materia.nombre,
            'grupo': grupo_codigo,
            'dia_raw': dia_raw,
            'hora_inicio': a.get_hora_inicio_str(),
            'hora_fin': a.get_hora_fin_str(),
            'horas_totales': duracion_horas,
            'carrera': a.materia.carrera.codigo if a.materia.carrera else ''
        }
        datos_profesor['clases'].append(detalle_clase)

    orden_dias = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado']
    datos_profesor['clases'].sort(
        key=lambda c: (orden_dias.index(c['dia_raw']) if c['dia_raw'] in orden_dias else 99, c['hora_inicio'])
    )

    periodo, año = obtener_periodo_actual()
    buffer = generar_excel_formato_fda(datos_profesor, periodo=periodo, año=año)

    filename = f"Horario_{profesor.nombre.replace(' ','_')}_{profesor.apellido.replace(' ','_')}.xlsx"
    return buffer, filename


def _generar_excel_horario_profesor(profesor_nombre):
    """
    Genera el archivo Excel del horario del profesor con formato de plantilla FDA (Carga Horaria).
    Utiliza generar_excel_profesor_buffer() y envuelve con send_file().
    """
    from flask import send_file

    try:
        buffer, filename = generar_excel_profesor_buffer(profesor_nombre)
        if buffer is None:
            return filename, 404

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Ocurrió un error al generar el archivo Excel: {e}", 500


def generar_pdf_profesor_buffer(profesor_nombre):
    """
    Genera un BytesIO con el PDF del horario del profesor en formato FDA.
    Retorna (BytesIO, filename) o (None, error_message).
    """
    from models import User, HorarioAcademico, ConfiguracionSistema

    try:
        profesor = User.query.filter(
            (User.nombre + ' ' + User.apellido) == profesor_nombre
        ).first()

        if not profesor:
            return None, "Profesor no encontrado"

        asignaciones = HorarioAcademico.query.filter_by(profesor_id=profesor.id, activo=True).all()

        periodo, anio = obtener_periodo_actual()
        fecha_inicio = ConfiguracionSistema.get_config('fecha_inicio_periodo', '') or ''

        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
        dias_map = {'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles',
                    'miércoles': 'Miércoles', 'jueves': 'Jueves', 'viernes': 'Viernes',
                    'sabado': 'Sábado', 'sábado': 'Sábado'}

        es_tc = profesor.is_profesor_completo()
        es_asignatura = profesor.is_profesor_asignatura()

        horas_imparticion = 0
        for a in asignaciones:
            if a.horario:
                duracion = (a.horario.hora_fin.hour - a.horario.hora_inicio.hour) + \
                           (a.horario.hora_fin.minute - a.horario.hora_inicio.minute) / 60.0
                horas_imparticion += duracion
        horas_imparticion = int(horas_imparticion) if horas_imparticion == int(horas_imparticion) else horas_imparticion

        # Horas TC: contar desde actividades PTC reales si existen, sino desde configuración
        from models import ActividadPTC as ActPTC
        ptc_real = ActPTC.query.filter_by(profesor_id=profesor.id, activo=True).all()
        if ptc_real:
            horas_asesoria = sum(1 for p in ptc_real if p.tipo_actividad == 'asesoria')
            horas_tutoria = sum(1 for p in ptc_real if p.tipo_actividad == 'tutoria')
            horas_gestion = sum(1 for p in ptc_real if p.tipo_actividad == 'gestion')
        else:
            horas_asesoria = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor.id}_asesoria', 0) or 0)
            horas_tutoria = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor.id}_tutoria', 0) or 0)
            horas_gestion = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor.id}_gestion', 0) or 0)
        horas_dual = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor.id}_dual', 0) or 0)
        horas_investigacion = int(ConfiguracionSistema.get_config(f'horas_tc_{profesor.id}_investigacion', 0) or 0)
        total_horas = horas_imparticion + horas_asesoria + horas_tutoria + horas_gestion + horas_dual + horas_investigacion

        # Nombres desde configuración
        nombre_director = ConfiguracionSistema.get_config('director_academico_nombre', 'Director Académico') or 'Director Académico'
        nombre_responsable = ConfiguracionSistema.get_config('responsable_pa_nombre', 'Responsable del PA') or 'Responsable del PA'

        # Firmas
        def cargar_firma_pdf(ruta):
            """Carga una imagen de firma para PDF, convirtiendo .webp si necesario"""
            if not ruta or not os.path.exists(ruta):
                return None
            try:
                ext = os.path.splitext(ruta)[1].lower()
                if ext == '.webp':
                    pil_img = PILImage.open(ruta)
                    if pil_img.mode in ('RGBA', 'LA', 'P'):
                        pil_img = pil_img.convert('RGBA')
                    else:
                        pil_img = pil_img.convert('RGB')
                    img_buffer = BytesIO()
                    pil_img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    return RlImage(img_buffer, width=1.3*inch, height=0.5*inch)
                else:
                    return RlImage(ruta, width=1.3*inch, height=0.5*inch)
            except Exception as e:
                logger.error(f"Error cargando firma PDF: {e}")
                return None

        # Firma del profesor
        firma_prof_elem = None
        if profesor.firma:
            firma_prof_elem = cargar_firma_pdf(os.path.join('static', 'uploads', 'firmas', profesor.firma))

        # Firma del director
        firma_dir_elem = None
        config_firma_dir = ConfiguracionSistema.query.filter_by(clave='director_academico_firma').first()
        if config_firma_dir and config_firma_dir.valor:
            firma_dir_elem = cargar_firma_pdf(os.path.join('static', 'uploads', 'firmas', config_firma_dir.valor))

        # Firma del responsable
        firma_resp_elem = None
        config_firma_resp = ConfiguracionSistema.query.filter_by(clave='responsable_pa_firma').first()
        if config_firma_resp and config_firma_resp.valor:
            firma_resp_elem = cargar_firma_pdf(os.path.join('static', 'uploads', 'firmas', config_firma_resp.valor))

        # =====================================================================
        # 2. CREAR PDF
        # =====================================================================
        buffer = BytesIO()
        # Portrait orientation (vertical) — letter = 612x792 pt
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=20, bottomMargin=20, leftMargin=24, rightMargin=24)
        elements = []
        styles = getSampleStyleSheet()

        verde_uptex_color = colors.Color(0, 0.69, 0.31)  # #00B050
        gris_claro = colors.Color(0.85, 0.85, 0.85)

        # Estilos personalizados
        title_style = ParagraphStyle('FDATitle', parent=styles['Heading1'], fontSize=12, alignment=1, spaceAfter=0, spaceBefore=0, fontName='Helvetica-Bold', textColor=colors.white)
        label_style = ParagraphStyle('FDALabel', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=1)
        label_white = ParagraphStyle('FDALabelW', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=1, textColor=colors.white)
        normal_style = ParagraphStyle('FDANormal', parent=styles['Normal'], fontSize=8, fontName='Helvetica', alignment=1)
        cell_style = ParagraphStyle('FDACell', parent=styles['Normal'], fontSize=5.5, leading=6.5, alignment=1, fontName='Helvetica')
        ptc_cell_style = ParagraphStyle('FDAPTCCell', parent=styles['Normal'], fontSize=5.5, leading=6.5, alignment=1, fontName='Helvetica-Bold', textColor=colors.white)
        ptc_tuto_cell_style = ParagraphStyle('FDATutoCell', parent=styles['Normal'], fontSize=5.5, leading=6.5, alignment=1, fontName='Helvetica-Bold', textColor=colors.HexColor('#212529'))
        small_style = ParagraphStyle('FDASmall', parent=styles['Normal'], fontSize=6, fontName='Helvetica', alignment=1)
        left_style = ParagraphStyle('FDALeft', parent=styles['Normal'], fontSize=8, fontName='Helvetica', alignment=0)
        note_style = ParagraphStyle('FDANote', parent=styles['Normal'], fontSize=6, fontName='Helvetica-Oblique', alignment=0, spaceAfter=0)
        bold_left = ParagraphStyle('BoldLeft', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=0)

        # Cargar logo
        logo_path = os.path.join('static', 'images', 'logo.png')
        logo_elem = None
        if os.path.exists(logo_path):
            try:
                logo_elem = RlImage(logo_path, width=0.65*inch, height=0.65*inch)
            except Exception:
                logo_elem = None

        # =====================================================================
        # ENCABEZADO - Logo + Carga Horaria + Área/Vigencia/Código (verde)
        # =====================================================================
        page_usable = 612 - 48  # letter width minus margins
        header_data = [
            [logo_elem or '', Paragraph('Carga Horaria', title_style), Paragraph('<b>Código:</b> FDA-02.5', label_white)],
            ['', Paragraph('<b>Área:</b> Dirección Académica', label_white), Paragraph(f'<b>Vigencia:</b> {periodo} {anio}', label_white)],
        ]
        header_table = Table(header_data, colWidths=[0.6*inch, 5.3*inch, 1.95*inch])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), verde_uptex_color),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('SPAN', (0, 0), (0, 1)),
            ('GRID', (0, 0), (-1, -1), 0.5, verde_uptex_color),
            ('LINEBELOW', (1, 0), (2, 0), 0.5, colors.white),
            ('ROWHEIGHTS', (0, 0), (-1, 0), 24),
            ('ROWHEIGHTS', (0, 1), (-1, 1), 16),
        ]))
        elements.append(header_table)

        # Nombre del profesor
        nombre_row = [
            Paragraph(f'<b>{profesor_nombre.upper()}</b>', ParagraphStyle('Name', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold', alignment=0)),
        ]
        nombre_table = Table([nombre_row], colWidths=[page_usable])
        nombre_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWHEIGHTS', (0, 0), (-1, -1), 16),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(nombre_table)

        # Tipo: Prof. Asignatura / Prof. Tiempo Completo (verde)
        tipo_row = [
            Paragraph(f"Prof. Asignatura: <b>{'X' if es_asignatura else '  '}</b>", label_white),
            Paragraph(f"Prof. Tiempo Completo: <b>{'X' if es_tc else '  '}</b>", label_white),
        ]
        tipo_table = Table([tipo_row], colWidths=[page_usable/2, page_usable/2])
        tipo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), verde_uptex_color),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.white),
            ('ROWHEIGHTS', (0, 0), (-1, -1), 14),
        ]))
        elements.append(tipo_table)

        # Periodo | Fecha de Inicio | Plan de Estudios
        periodo_row = [
            Paragraph(f'<b>Periodo:</b> {periodo}', bold_left),
            Paragraph(f'<b>Fecha de Inicio:</b> {fecha_inicio}', label_style),
            Paragraph(f'<b>Plan de Estudios:</b> {anio}', label_style),
        ]
        periodo_table = Table([periodo_row], colWidths=[page_usable/3]*3)
        periodo_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWHEIGHTS', (0, 0), (-1, -1), 14),
        ]))
        elements.append(periodo_table)

        elements.append(Paragraph('<b>Instrucciones:</b> Introducir nombre de la Asignatura, Salón y Grupo dentro de la celda correspondiente al día y la hora que será impartida.', ParagraphStyle('Inst', parent=styles['Normal'], fontSize=6, fontName='Helvetica', spaceAfter=2, spaceBefore=2)))

        # =====================================================================
        # CUADRÍCULA DE HORARIOS (encabezados verdes)
        # =====================================================================
        # Query PTC activities for this professor
        from models import Horario as HorarioModel, ActividadPTC
        ptc_acts = []
        if es_tc:
            ptc_acts = ActividadPTC.query.filter_by(
                profesor_id=profesor.id, activo=True
            ).all()
        ptc_lookup = {}
        for p in ptc_acts:
            if p.horario and p.horario.hora_inicio:
                ptc_lookup[(p.horario.hora_inicio.hour, p.dia_semana.lower())] = p
        ptc_labels_map = {'asesoria': 'ASESORÍA', 'tutoria': 'TUTORÍA', 'gestion': 'GESTIÓN'}
        COLOR_ASESORIA = colors.HexColor('#28A745')
        COLOR_TUTORIA = colors.HexColor('#FFC107')
        COLOR_GESTION = colors.HexColor('#DC3545')
        ptc_bg_map = {'asesoria': COLOR_ASESORIA, 'tutoria': COLOR_TUTORIA, 'gestion': COLOR_GESTION}

        # Use only days L-V (no Sábado unless needed)
        dias_semana_pdf = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
        dias_rev_pdf = {'Lunes': 'lunes', 'Martes': 'martes', 'Miércoles': 'miercoles', 'Jueves': 'jueves', 'Viernes': 'viernes', 'Sábado': 'sabado'}
        tiene_sabado = any(a.dia_semana.lower() in ('sabado','sábado') for a in asignaciones) or any(p.dia_semana.lower() in ('sabado','sábado') for p in ptc_acts)
        if tiene_sabado:
            dias_semana_pdf.append('Sábado')

        grid_data = [[Paragraph('<b>Hora</b>', label_white)] + [Paragraph(f'<b>{d}</b>', label_white) for d in dias_semana_pdf]]

        horarios_sistema_pdf = HorarioModel.query.filter_by(activo=True).order_by(HorarioModel.hora_inicio).all()
        horas_pdf = []
        seen_pdf = set()
        for hs in horarios_sistema_pdf:
            h_hour = hs.hora_inicio.hour
            if h_hour not in seen_pdf:
                seen_pdf.add(h_hour)
                horas_pdf.append((h_hour, hs.hora_inicio, hs.hora_fin))

        # Build clase lookup
        clase_lk = {}
        for a in asignaciones:
            if a.horario and a.horario.hora_inicio:
                dia_key = dias_map.get(a.dia_semana.lower(), '')
                if dia_key:
                    clase_lk[(a.horario.hora_inicio.hour, dias_rev_pdf.get(dia_key, a.dia_semana.lower()))] = a

        ptc_bg_cmds = []  # (col, row, color) for PTC cell backgrounds

        for hora, h_inicio, h_fin in horas_pdf:
            # Skip hours with no content at all
            has_content = False
            for dia in dias_semana_pdf:
                db_dia = dias_rev_pdf.get(dia, dia.lower())
                if (hora, db_dia) in clase_lk or (hora, db_dia) in ptc_lookup:
                    has_content = True
                    break
            if not has_content:
                continue

            hora_str = f"{h_inicio.strftime('%H:%M')}\n{h_fin.strftime('%H:%M')}"
            row = [Paragraph(f'<b>{h_inicio.strftime("%H:%M")}<br/>{h_fin.strftime("%H:%M")}</b>', label_style)]

            for cidx, dia in enumerate(dias_semana_pdf):
                db_dia = dias_rev_pdf.get(dia, dia.lower())
                ck = (hora, db_dia)
                if ck in clase_lk:
                    a = clase_lk[ck]
                    grupo_codigo = a.grupo if a.grupo else ''
                    row.append(Paragraph(f'{a.materia.nombre} {grupo_codigo}', cell_style))
                elif ck in ptc_lookup:
                    p = ptc_lookup[ck]
                    lbl = ptc_labels_map.get(p.tipo_actividad, p.tipo_actividad.upper())
                    st = ptc_tuto_cell_style if p.tipo_actividad == 'tutoria' else ptc_cell_style
                    row.append(Paragraph(f'{lbl}', st))
                    ptc_bg_cmds.append((cidx + 1, len(grid_data), ptc_bg_map.get(p.tipo_actividad, colors.grey)))
                else:
                    row.append('')

            grid_data.append(row)

        n_dias = len(dias_semana_pdf)
        hora_col_w = 0.55 * inch
        dia_col_w = (page_usable - hora_col_w) / n_dias
        col_widths = [hora_col_w] + [dia_col_w] * n_dias
        grid_table = Table(grid_data, colWidths=col_widths)

        grid_style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), verde_uptex_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BACKGROUND', (0, 1), (0, -1), colors.Color(0.95, 0.95, 0.95)),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.75, 0.75, 0.75)),
            ('ROWHEIGHTS', (0, 0), (-1, -1), 18),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ]
        # Add PTC cell background colors
        for (col, row_idx, bg) in ptc_bg_cmds:
            grid_style_cmds.append(('BACKGROUND', (col, row_idx), (col, row_idx), bg))

        grid_table.setStyle(TableStyle(grid_style_cmds))
        elements.append(grid_table)

        elements.append(Spacer(1, 6))

        # =====================================================================
        # TABLA DE TIPOS DE HORAS (encabezados verdes)
        # =====================================================================
        tipos_horas_pdf = [
            ('Impartición de Curso', horas_imparticion),
            ('Asesoría', horas_asesoria),
            ('Tutoría', horas_tutoria),
            ('Apoyo a la Gestión', horas_gestion),
            ('Dual', horas_dual),
            ('Investigación', horas_investigacion),
        ]

        horas_data = [
            [Paragraph('<b>Tipo de Horas</b>', label_white), Paragraph('<b>Horas</b>', label_white)],
        ]
        for tipo, val in tipos_horas_pdf:
            display_val = '' if val == 0 or val == '' else str(val)
            horas_data.append([Paragraph(tipo, left_style), Paragraph(display_val, normal_style)])

        horas_data.append([
            Paragraph('<b>Total de Horas</b>', ParagraphStyle('R', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold', alignment=2)),
            Paragraph(f'<b>{total_horas}</b>', label_style)
        ])

        horas_table = Table(horas_data, colWidths=[2.0*inch, 0.8*inch])
        horas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), verde_uptex_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.75, 0.75, 0.75)),
            ('ROWHEIGHTS', (0, 0), (-1, -1), 12),
        ]))
        elements.append(horas_table)
        elements.append(Paragraph('*Solo llenar en caso de ser Profesor de Tiempo Completo', note_style))
        elements.append(Spacer(1, 4))

        # =====================================================================
        # SECCIÓN DE FIRMAS
        # =====================================================================
        cargo_profesor = "PROFESOR DE TIEMPO COMPLETO" if es_tc else "PROFESOR DE ASIGNATURA"

        firma_data = [
            [Paragraph('<b>Elaboró:</b>', label_style),
             Paragraph('<b>Autorizó:</b>', label_style),
             Paragraph('<b>Recibió:</b>', label_style)],
            [Paragraph(f'<b>{profesor_nombre.upper()}</b>', label_style),
             Paragraph(f'<b>{nombre_director}</b>', label_style),
             Paragraph(f'<b>{nombre_responsable}</b>', label_style)],
            [Paragraph(cargo_profesor, small_style),
             Paragraph('Director Académico', small_style),
             Paragraph('Responsable del PA', small_style)],
            [firma_prof_elem or '',
             firma_dir_elem or '',
             firma_resp_elem or ''],
            [Paragraph('Firma', label_style),
             Paragraph('Firma', label_style),
             Paragraph('Firma', label_style)],
        ]

        fw = page_usable / 3
        firma_table = Table(firma_data, colWidths=[fw, fw, fw])
        firma_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), gris_claro),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (0, -1), 0.5, colors.grey),
            ('BOX', (1, 0), (1, -1), 0.5, colors.grey),
            ('BOX', (2, 0), (2, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.grey),
            ('LINEBELOW', (0, -1), (-1, -1), 0.5, colors.grey),
            ('ROWHEIGHTS', (0, 3), (-1, 3), 40),
        ]))
        elements.append(firma_table)

        # =====================================================================
        # 3. GUARDAR EN BUFFER
        # =====================================================================
        doc.build(elements)
        buffer.seek(0)

        filename = f"Horario_{profesor.nombre.replace(' ','_')}_{profesor.apellido.replace(' ','_')}.pdf"
        return buffer, filename

    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, f"Ocurrió un error al generar el archivo PDF: {e}"


def _generar_pdf_horario_profesor(profesor_nombre):
    """
    Genera el archivo PDF del horario del profesor con formato de plantilla FDA.
    Utiliza generar_pdf_profesor_buffer() y envuelve con send_file().
    """
    from flask import send_file

    try:
        buffer, filename = generar_pdf_profesor_buffer(profesor_nombre)
        if buffer is None:
            return filename, 500

        return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Ocurrió un error al generar el archivo PDF: {e}", 500
